from __future__ import annotations

import asyncio
import base64
import json
import os
import re
import shutil
import subprocess
import tempfile
import textwrap
import time
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import urlparse
from urllib.request import Request, urlopen

from aetherion_sdk import tool
from common_lib.storage.storage_client import RetrievalMode, storage
from common_lib.utils.logger import setup_logger
from src.runtime.parameterization import (
    normalize_param_name as _normalize_param_name,
    parameterise_script as _parameterise_script,
    substitute_parameters as _substitute_parameters,
)
from src.utils.html_report_generator import generate_html_report_content

logger = setup_logger(__name__)

_MAX_AI_LOG_CHARS = 3_000


_MAX_AI_IMAGE_BYTES = 700_000
_MAX_AI_STEP_IMAGES = 2
_MAX_OPENAI_ERROR_CHARS = 1_600
_MAX_FLOW_CONTEXT_PAGE_TEXT_CHARS = 5_000
_RUNNER_CONFIG_LINE_RE = re.compile(
    r"^\s*(?:export\s+)?(?P<key>[A-Za-z_][A-Za-z0-9_]*)\s*=\s*(?P<value>.*?)\s*$"
)
_FLOW_CONTEXT_SHEET_ALIASES = {
    "flow_context",
    "flowcontext",
    "flow_io",
    "flowio",
    "context_io",
    "contextio",
    "input_output",
    "inputoutput",
    "output_input",
    "outputinput",
}
_PLACEHOLDER_TOKEN_RE = re.compile(r"\{\{([A-Za-z0-9_]+)\}\}")
_RUNNER_PROJECT_ROOT = Path(__file__).resolve().parents[2]
_RUNNER_CONFIG_PATH = _RUNNER_PROJECT_ROOT / "configs.txt"
_RUNNER_DATA_DIR = _RUNNER_PROJECT_ROOT / ".runner_data"


def _get_bucket_name() -> str:
    bucket_name = os.getenv("STORAGE_ACTIVITIES_BUCKET", "").strip()
    if not bucket_name:
        raise RuntimeError("STORAGE_ACTIVITIES_BUCKET is not configured.")
    return bucket_name


def _safe_segment(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", (value or "").strip())
    return cleaned.strip("._") or "unknown"


def _load_runner_env_defaults(config_path: Path | None = None) -> dict[str, str]:
    path = config_path or _RUNNER_CONFIG_PATH
    if not path.exists():
        return {}

    try:
        raw_lines = path.read_text(encoding="utf-8").splitlines()
    except Exception:
        logger.exception("Failed to read runner config defaults from %s", path)
        return {}

    defaults: dict[str, str] = {}
    for raw_line in raw_lines:
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        match = _RUNNER_CONFIG_LINE_RE.match(line)
        if not match:
            continue
        value = match.group("value").strip()
        if len(value) >= 2 and value[0] == value[-1] and value[0] in {"'", '"'}:
            value = value[1:-1]
        defaults[match.group("key")] = value
    return defaults


def _merge_runner_env_defaults(
    env: dict[str, str],
    *,
    config_path: Path | None = None,
) -> dict[str, str]:
    merged = dict(env)
    for key, value in _load_runner_env_defaults(config_path).items():
        merged.setdefault(key, value)
    return merged


def _ensure_runner_pythonpath(env: dict[str, str], project_root: Path | None = None) -> dict[str, str]:
    updated = dict(env)
    root = str((project_root or _RUNNER_PROJECT_ROOT).resolve())
    existing = str(updated.get("PYTHONPATH") or "").strip()
    entries = [entry for entry in existing.split(os.pathsep) if entry] if existing else []
    if root not in entries:
        entries.insert(0, root)
    updated["PYTHONPATH"] = os.pathsep.join(entries)
    return updated


def _default_experience_store_path(project_root: Path | None = None) -> Path:
    root = (project_root or _RUNNER_PROJECT_ROOT).resolve()
    return root / ".runner_data" / "experience.jsonl"


def _env_flag(value: str | None, default: bool) -> bool:
    raw_value = "" if value is None else str(value).strip()
    if not raw_value:
        return default
    return raw_value.lower() not in ("false", "0", "no", "off")


def _split_storage_object_ref(object_ref: str) -> tuple[str, str]:
    raw = str(object_ref or "").strip()
    if not raw:
        raise ValueError("Storage object key is required.")

    if raw.lower().startswith("s3://"):
        parsed = urlparse(raw)
        bucket_name = parsed.netloc.strip()
        object_key = parsed.path.lstrip("/")
        if not bucket_name or not object_key:
            raise ValueError(f"Invalid S3 object reference: {object_ref}")
        return bucket_name, object_key

    bucket_name = _get_bucket_name()
    bucket_prefix = f"{bucket_name}/"
    object_key = raw[len(bucket_prefix) :] if raw.startswith(bucket_prefix) else raw
    object_key = object_key.lstrip("/")
    if not object_key:
        raise ValueError("Storage object key is required.")
    return bucket_name, object_key


def _storage_get_bytes(object_key: str) -> bytes:
    storage.init_client()
    bucket_name, normalized_key = _split_storage_object_ref(object_key)

    if hasattr(storage, "retrieve"):
        data = storage.retrieve(
            bucket_name=bucket_name,
            object_key=normalized_key,
            retrieval_mode=RetrievalMode.FULL_OBJECT,
        )
        if isinstance(data, bytes):
            return data

    client = getattr(storage, "client", None)
    if client is None:
        raise RuntimeError("Storage client is not initialized.")

    response = client.get_object(Bucket=bucket_name, Key=normalized_key)
    return response["Body"].read()


def _storage_put_bytes(object_key: str, data: bytes, *, content_type: str) -> tuple[str, str]:
    storage.init_client()
    return storage.store_object(
        bucket_name=_get_bucket_name(),
        object_key=object_key,
        data=data,
        content_type=content_type,
    )


def _load_script_bytes(object_key: str) -> bytes:
    if not object_key.lower().endswith(".py"):
        raise ValueError(f"Unsupported recording artifact: {object_key}. playwright_test_runner can only execute .py recordings.")
    return _storage_get_bytes(object_key)


def _read_manifest(object_key: str) -> dict[str, Any]:
    manifest_bytes = _storage_get_bytes(object_key)
    return json.loads(manifest_bytes.decode("utf-8"))


def _is_missing_storage_object_error(exc: Exception) -> bool:
    if isinstance(exc, FileNotFoundError):
        return True

    response = getattr(exc, "response", None)
    if isinstance(response, dict):
        error_code = str(response.get("Error", {}).get("Code") or "").strip()
        if error_code in {"404", "NoSuchKey", "NotFound"}:
            return True

    message = str(exc).lower()
    return any(
        token in message
        for token in (
            "nosuchkey",
            "no such key",
            "not found",
            "404",
        )
    )


def _recording_artifact_identity(recording: dict[str, Any]) -> str:
    recording_id = str(recording.get("id") or "unknown")
    file_key = str(recording.get("file") or recording.get("recording_name") or "").strip()
    recording_name = str(recording.get("name") or "").strip() or file_key or recording_id
    parameter_row_index = recording.get("parameter_row_index")
    parameter_set_index = recording.get("parameter_set_index")

    artifact_identity = file_key or recording_name or recording_id
    if parameter_row_index not in (None, ""):
        artifact_identity = f"{artifact_identity}__row_{parameter_row_index}"
    elif parameter_set_index not in (None, ""):
        artifact_identity = f"{artifact_identity}__set_{parameter_set_index}"
    return artifact_identity


def _manifest_key_for_recording(
    test_suite_id: str,
    run_id: str,
    recording: dict[str, Any],
) -> str:
    artifact_identity = _recording_artifact_identity(recording)
    artifact_prefix = (
        f"playwright-test-results/{_safe_segment(test_suite_id)}/{_safe_segment(run_id)}"
        f"/{_safe_segment(artifact_identity)}"
    )
    return f"{artifact_prefix}/manifest.json"


def _load_resume_state_from_run_data(
    test_suite_id: str,
    previous_run_id: str,
    recordings: list[dict[str, Any]],
) -> dict[str, Any]:
    if not previous_run_id:
        raise ValueError("previous_run_id is required")

    previous_results: list[dict[str, Any]] = []
    failed_index: int | None = None
    failed_recording_name = ""

    for idx, recording in enumerate(recordings):
        manifest_key = _manifest_key_for_recording(test_suite_id, previous_run_id, recording)
        display_name = str(recording.get("name") or recording.get("file") or f"recording-{idx}")
        try:
            manifest = _read_manifest(manifest_key)
        except Exception as exc:
            if _is_missing_storage_object_error(exc):
                failed_index = idx
                failed_recording_name = display_name
                break
            raise

        status = str(manifest.get("status") or "").strip().lower()
        if status == "passed":
            previous_results.append(manifest)
            continue

        failed_index = idx
        failed_recording_name = str(manifest.get("recording_name") or display_name)
        break

    if failed_index is None:
        raise ValueError(f"No failed recording found in previous run {previous_run_id}")

    return {
        "resume_start_index": failed_index,
        "previous_results": previous_results,
        "failed_recording_name": failed_recording_name,
    }


def _extract_table_parameter_sets(rows: list[tuple[Any, ...]]) -> list[dict[str, Any]]:
    normalized_rows: list[tuple[int, tuple[str, ...]]] = []
    for row_index, row in enumerate(rows, start=1):
        values = tuple(str(value if value is not None else "").strip() for value in row)
        if any(values):
            normalized_rows.append((row_index, values))

    if not normalized_rows:
        return []

    header_aliases = {
        "parameter",
        "parameter name",
        "param",
        "name",
        "field",
    }
    value_aliases = {
        "value",
        "parameter value",
        "default",
        "default value",
    }

    def _header_index(cells: tuple[str, ...], aliases: set[str]) -> int | None:
        for idx, cell in enumerate(cells):
            normalized = re.sub(r"\s+", " ", cell.lower()).strip()
            if normalized in aliases:
                return idx
        return None

    def _build_vertical_parameter_set(
        *,
        start_row: int,
        param_idx: int,
        value_idx: int,
    ) -> list[dict[str, Any]]:
        params: dict[str, str] = {}
        first_data_row_index = first_row_index
        for source_row_index, row in normalized_rows[start_row:]:
            if param_idx >= len(row) or value_idx >= len(row):
                continue
            param_name = row[param_idx]
            param_value = row[value_idx]
            # Skip action rows (click_*), header-like rows, and rows with no value
            if not param_name or param_name.lower().startswith("click_") or not param_value:
                continue
            if not params:
                first_data_row_index = source_row_index
            params[_normalize_param_name(param_name)] = param_value
        if not params:
            return []
        return [
            {
                "row_index": first_data_row_index,
                "values": params,
            }
        ]

    first_row_index, first_row = normalized_rows[0]
    header_positions: dict[str, set[int]] = {}
    for idx, header in enumerate(first_row):
        normalized_header = _normalize_param_name(header)
        if not normalized_header:
            continue
        header_positions.setdefault(normalized_header, set()).add(idx)

    def _looks_like_misaligned_header_entry(row: tuple[str, ...]) -> bool:
        non_empty_cells = [
            (idx, _normalize_param_name(value))
            for idx, value in enumerate(row)
            if str(value or "").strip()
        ]
        if len(non_empty_cells) > 2:
            return False
        for cell_index, normalized_value in non_empty_cells:
            positions = header_positions.get(normalized_value) or set()
            if positions and cell_index not in positions:
                return True
        return False

    param_idx = _header_index(first_row, header_aliases)
    value_idx = _header_index(first_row, value_aliases)
    has_explicit_parameter_header = param_idx is not None and value_idx is not None
    if has_explicit_parameter_header:
        parameter_sets = _build_vertical_parameter_set(
            start_row=1,
            param_idx=param_idx,
            value_idx=value_idx,
        )
        if parameter_sets:
            return parameter_sets

    if len(normalized_rows) >= 2 and sum(1 for cell in first_row if cell) > 2:
        parameter_sets: list[dict[str, Any]] = []
        for source_row_index, data_row in normalized_rows[1:]:
            if _looks_like_misaligned_header_entry(data_row):
                continue
            horizontal_params: dict[str, str] = {}
            for idx, header in enumerate(first_row):
                if not header or idx >= len(data_row):
                    continue
                value = data_row[idx]
                if not value:
                    continue
                normalized_header = _normalize_param_name(header)
                if normalized_header.startswith("click_"):
                    continue
                horizontal_params[normalized_header] = value
            if horizontal_params:
                parameter_sets.append(
                    {
                        "row_index": source_row_index,
                        "values": horizontal_params,
                    }
                )
        if parameter_sets:
            return parameter_sets

    if param_idx is None:
        param_idx = 0
    if value_idx is None:
        value_idx = 1

    return _build_vertical_parameter_set(
        start_row=0,
        param_idx=param_idx,
        value_idx=value_idx,
    )


def _extract_table_parameters(rows: list[tuple[Any, ...]]) -> dict[str, str]:
    parameter_sets = _extract_table_parameter_sets(rows)
    if not parameter_sets:
        return {}
    return dict(parameter_sets[0].get("values") or {})


def _parse_excel_parameter_sets(raw_bytes: bytes) -> list[dict[str, Any]]:
    import io
    import openpyxl  # lazy import — only needed when a parameters file is provided

    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    ws = wb.active
    parameter_sets = _extract_table_parameter_sets(list(ws.iter_rows(values_only=True)))
    wb.close()
    return parameter_sets


def _parse_excel_parameters(raw_bytes: bytes) -> dict[str, str]:
    parameter_sets = _parse_excel_parameter_sets(raw_bytes)
    if not parameter_sets:
        return {}
    return dict(parameter_sets[0].get("values") or {})


def _coerce_flow_context_bool(value: Any, *, default: bool) -> bool:
    raw = str(value or "").strip().lower()
    if not raw:
        return default
    return raw not in {"0", "false", "no", "off"}


def _coerce_flow_context_int(value: Any, *, default: int | None) -> int | None:
    raw = str(value or "").strip()
    if not raw:
        return default
    try:
        return int(raw)
    except (TypeError, ValueError):
        return default


def _parse_flow_context_aliases(value: Any) -> list[str]:
    if isinstance(value, (list, tuple, set)):
        aliases: list[str] = []
        seen: set[str] = set()
        for item in value:
            for alias in _parse_flow_context_aliases(item):
                normalized = alias.lower()
                if normalized in seen:
                    continue
                seen.add(normalized)
                aliases.append(alias)
        return aliases

    raw = str(value or "").strip()
    if not raw:
        return []
    parts = re.split(r"[|,\n;\r]+", raw)
    aliases: list[str] = []
    seen: set[str] = set()
    for part in parts:
        cleaned = re.sub(r"\s+", " ", str(part or "").strip())
        if not cleaned:
            continue
        normalized = cleaned.lower()
        if normalized in seen:
            continue
        seen.add(normalized)
        aliases.append(cleaned)
    return aliases


def _extract_flow_context_sheet_specs(rows: list[tuple[Any, ...]]) -> list[dict[str, Any]]:
    normalized_rows: list[tuple[int, tuple[str, ...]]] = []
    for row_index, row in enumerate(rows, start=1):
        values = tuple(str(value if value is not None else "").strip() for value in row)
        if any(values):
            normalized_rows.append((row_index, values))

    if len(normalized_rows) < 2:
        return []

    _, header_row = normalized_rows[0]
    header_map: dict[str, int] = {}
    for index, cell in enumerate(header_row):
        normalized = _normalize_param_name(cell)
        if normalized:
            header_map.setdefault(normalized, index)

    def _row_value(row: tuple[str, ...], *aliases: str) -> str:
        for alias in aliases:
            idx = header_map.get(alias)
            if idx is None or idx >= len(row):
                continue
            value = row[idx]
            if value:
                return value
        return ""

    specs: list[dict[str, Any]] = []
    for row_index, row in normalized_rows[1:]:
        kind = _row_value(row, "kind", "direction", "type").lower() or "output"
        if kind not in {"input", "output"}:
            continue
        name = _normalize_param_name(_row_value(row, "name", "field", "key"))
        label = _row_value(row, "label", "title", "display_name")
        source = _row_value(row, "source", "extract_from", "surface").lower() or "auto"
        if not name and not label:
            continue
        if not name:
            name = _normalize_param_name(label)
        if not name:
            continue
        row_value = _row_value(row, "row")
        table_value = _row_value(row, "table_index", "table")
        specs.append(
            {
                "row_index": row_index,
                "kind": kind,
                "name": name,
                "label": label,
                "aliases": _parse_flow_context_aliases(
                    _row_value(row, "aliases", "alias", "alternate_labels", "alternate_label")
                ),
                "source": source,
                "pattern": _row_value(row, "pattern", "regex"),
                "group": _coerce_flow_context_int(_row_value(row, "group"), default=1),
                "column": _row_value(row, "column", "header"),
                "row": _coerce_flow_context_int(row_value, default=None if not row_value else 0),
                "table_index": _coerce_flow_context_int(table_value, default=None if not table_value else 0),
                "required": _coerce_flow_context_bool(_row_value(row, "required", "is_required"), default=(kind == "output")),
                "prompt": _row_value(row, "prompt", "hint", "description", "instructions"),
                "value_type": _row_value(row, "value_type", "datatype", "data_type").lower() or "text",
                "use_ai": _coerce_flow_context_bool(_row_value(row, "use_ai", "allow_ai"), default=True),
            }
        )

    return specs


def _parse_excel_flow_context_specs(raw_bytes: bytes) -> list[dict[str, Any]]:
    import io
    import openpyxl  # lazy import — only needed when a workbook is provided

    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            if _normalize_param_name(ws.title) not in _FLOW_CONTEXT_SHEET_ALIASES:
                continue
            return _extract_flow_context_sheet_specs(list(ws.iter_rows(values_only=True)))
    finally:
        wb.close()
    return []


def _parse_csv_parameter_sets(raw_bytes: bytes) -> list[dict[str, Any]]:
    import csv
    import io

    reader = csv.reader(io.StringIO(raw_bytes.decode("utf-8-sig")))
    return _extract_table_parameter_sets([tuple(row) for row in reader])


def _parse_csv_parameters(raw_bytes: bytes) -> dict[str, str]:
    parameter_sets = _parse_csv_parameter_sets(raw_bytes)
    if not parameter_sets:
        return {}
    return dict(parameter_sets[0].get("values") or {})


def _load_parameter_sets_from_file(file_key: str) -> list[dict[str, Any]]:
    raw_bytes = _storage_get_bytes(file_key)
    lower = file_key.lower()
    if lower.endswith(".xlsx") or lower.endswith(".xls"):
        return _parse_excel_parameter_sets(raw_bytes)
    if lower.endswith(".csv"):
        return _parse_csv_parameter_sets(raw_bytes)
    raise ValueError(f"Unsupported parameters file format: {file_key}. Use .xlsx or .csv.")


def _load_parameters_from_file(file_key: str) -> dict[str, str]:
    parameter_sets = _load_parameter_sets_from_file(file_key)
    if not parameter_sets:
        return {}
    return dict(parameter_sets[0].get("values") or {})


def _load_flow_context_specs_from_file(file_key: str) -> list[dict[str, Any]]:
    raw_bytes = _storage_get_bytes(file_key)
    lower = file_key.lower()
    if lower.endswith(".xlsx") or lower.endswith(".xls"):
        return _parse_excel_flow_context_specs(raw_bytes)
    return []


def _derive_parameters_file_candidates(file_key: str) -> list[str]:
    raw = str(file_key or "").strip()
    if not raw:
        return []

    bucket_name, normalized_key = _split_storage_object_ref(raw)
    script_path = Path(normalized_key)
    if script_path.suffix.lower() != ".py":
        return []

    sibling_keys = [
        str(script_path.with_name(f"{script_path.stem}_params.xlsx")),
        str(script_path.with_name(f"{script_path.stem}_params.csv")),
    ]
    candidates: list[str] = []
    raw_uses_s3_uri = raw.lower().startswith("s3://")
    raw_uses_bucket_prefix = raw.startswith(f"{bucket_name}/")
    for key in sibling_keys:
        if raw_uses_s3_uri:
            candidates.append(f"s3://{bucket_name}/{key}")
        elif raw_uses_bucket_prefix:
            candidates.append(f"{bucket_name}/{key}")
        candidates.append(key)
    return list(dict.fromkeys(candidates))


def _load_recording_parameters(
    recording: dict[str, Any],
    file_key: str,
) -> tuple[dict[str, str], str | None]:
    parameter_sets, loaded_from = _load_recording_parameter_sets(recording, file_key)
    if not parameter_sets:
        return {}, loaded_from
    return dict(parameter_sets[0].get("values") or {}), loaded_from


def _load_recording_parameter_sets(
    recording: dict[str, Any],
    file_key: str,
) -> tuple[list[dict[str, Any]], str | None]:
    explicit_file = str(recording.get("parameters_file") or "").strip()
    candidates = []
    if explicit_file:
        candidates.append(explicit_file)
    candidates.extend(_derive_parameters_file_candidates(file_key))

    seen: set[str] = set()
    errors: list[tuple[str, Exception]] = []
    for candidate in candidates:
        if not candidate or candidate in seen:
            continue
        seen.add(candidate)
        try:
            return _load_parameter_sets_from_file(candidate), candidate
        except Exception as exc:
            errors.append((candidate, exc))

    if explicit_file:
        details = "; ".join(f"{candidate}: {exc}" for candidate, exc in errors)
        raise RuntimeError(f"Failed to load parameters file. {details}")

    return [], None


def _load_recording_flow_context_specs(
    recording: dict[str, Any],
    file_key: str,
) -> tuple[list[dict[str, Any]], str | None]:
    explicit_file = str(
        recording.get("flow_context_file")
        or recording.get("parameters_file_key")
        or recording.get("parameters_file")
        or ""
    ).strip()
    candidates: list[str] = []
    if explicit_file:
        candidates.append(explicit_file)
    candidates.extend(
        candidate
        for candidate in _derive_parameters_file_candidates(file_key)
        if candidate.lower().endswith((".xlsx", ".xls"))
    )

    seen: set[str] = set()
    for candidate in candidates:
        if not candidate or candidate in seen:
            continue
        seen.add(candidate)
        try:
            specs = _load_flow_context_specs_from_file(candidate)
            if specs:
                return specs, candidate
        except Exception as exc:
            logger.warning("Failed to load flow context specs from %s: %s", candidate, exc)
    return [], None


def _expand_recordings_for_parameter_rows_data(
    recordings: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    expanded_recordings: list[dict[str, Any]] = []

    for recording in recordings:
        if recording.get("skip_parameters_file_load"):
            expanded_recordings.append(recording)
            continue

        file_key = str(recording.get("file") or recording.get("recording_name") or "").strip()
        if not file_key:
            expanded_recordings.append(recording)
            continue

        try:
            parameter_sets, loaded_from = _load_recording_parameter_sets(recording, file_key)
        except Exception as exc:
            logger.warning("Failed to pre-expand parameters for %s: %s", file_key, exc)
            expanded_recordings.append(recording)
            continue

        if len(parameter_sets) <= 1:
            expanded_recordings.append(recording)
            continue

        base_name = str(recording.get("name") or file_key or "recording").strip() or "recording"
        base_parameters = recording.get("parameters") if isinstance(recording.get("parameters"), dict) else {}

        for set_index, parameter_set in enumerate(parameter_sets, start=1):
            row_index = int(parameter_set.get("row_index") or set_index)
            row_values = dict(parameter_set.get("values") or {})
            merged_parameters = dict(row_values)
            merged_parameters.update(base_parameters)

            expanded_recording = dict(recording)
            expanded_recording["id"] = f'{recording.get("id") or "recording"}-row-{row_index}'
            expanded_recording["name"] = f"{base_name} [row {row_index}]"
            expanded_recording["parameters"] = merged_parameters
            expanded_recording["parameters_file_key"] = loaded_from
            expanded_recording["parameter_set_index"] = set_index
            expanded_recording["parameter_row_index"] = row_index
            expanded_recording["skip_parameters_file_load"] = True
            expanded_recordings.append(expanded_recording)

    return expanded_recordings


def _normalize_parameter_values(parameters: dict[str, Any] | None) -> dict[str, str]:
    normalized: dict[str, str] = {}
    for raw_key, raw_value in (parameters or {}).items():
        param_name = _normalize_param_name(str(raw_key or ""))
        param_value = str(raw_value).strip() if raw_value is not None else ""
        if not param_name or not param_value:
            continue
        normalized[param_name] = param_value
    return normalized


def _parameters_to_json_object(parameters: dict[str, Any] | None) -> dict[str, str]:
    normalized = _resolve_parameter_references(parameters)
    return json.loads(json.dumps(normalized, sort_keys=True))


def _resolve_parameter_references(parameters: dict[str, Any] | None) -> dict[str, str]:
    resolved = _normalize_parameter_values(parameters)
    if not resolved:
        return {}

    for _ in range(len(resolved)):
        changed = False
        for key, value in list(resolved.items()):
            updated = _substitute_parameters(value, resolved)
            if updated == value:
                continue
            resolved[key] = updated
            changed = True
        if not changed:
            break
    return resolved


def _find_unresolved_placeholders(value: Any) -> list[str]:
    return [match.group(1) for match in _PLACEHOLDER_TOKEN_RE.finditer(str(value or ""))]


def _collect_unresolved_execution_parameters(parameters: dict[str, Any] | None) -> dict[str, list[str]]:
    unresolved: dict[str, list[str]] = {}
    for name, value in (parameters or {}).items():
        placeholders = _find_unresolved_placeholders(value)
        if not placeholders:
            continue
        unresolved[str(name)] = list(dict.fromkeys(placeholders))
    return unresolved


def _normalize_flow_context_specs(specs: Any) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    allowed_sources = {
        "auto",
        "page_semantics",
        "page_text",
        "page_url",
        "page_title",
        "stdout",
        "stderr",
        "oracle_table",
        "ai",
    }
    allowed_kinds = {"input", "output"}

    for item in specs or []:
        if not isinstance(item, dict):
            continue
        kind = str(item.get("kind") or "output").strip().lower() or "output"
        if kind not in allowed_kinds:
            continue
        name = _normalize_param_name(str(item.get("name") or ""))
        label = str(item.get("label") or "").strip()
        if not name and label:
            name = _normalize_param_name(label)
        if not name:
            continue
        source = str(item.get("source") or "auto").strip().lower() or "auto"
        if source not in allowed_sources:
            source = "auto"
        normalized.append(
            {
                "row_index": _coerce_flow_context_int(item.get("row_index"), default=None),
                "kind": kind,
                "name": name,
                "label": label,
                "aliases": _parse_flow_context_aliases(item.get("aliases")),
                "source": source,
                "pattern": str(item.get("pattern") or "").strip(),
                "group": _coerce_flow_context_int(item.get("group"), default=1),
                "column": str(item.get("column") or "").strip(),
                "row": _coerce_flow_context_int(item.get("row"), default=None),
                "table_index": _coerce_flow_context_int(item.get("table_index"), default=None),
                "required": _coerce_flow_context_bool(item.get("required"), default=(kind == "output")),
                "prompt": str(item.get("prompt") or "").strip(),
                "value_type": str(item.get("value_type") or "text").strip().lower() or "text",
                "use_ai": _coerce_flow_context_bool(item.get("use_ai"), default=True),
            }
        )
    return normalized


def _validate_flow_context_inputs(
    execution_parameters: dict[str, str],
    flow_context_specs: list[dict[str, Any]],
) -> tuple[list[str], dict[str, Any]]:
    input_specs = [spec for spec in flow_context_specs if spec.get("kind") == "input"]
    input_status: dict[str, Any] = {}
    missing: list[str] = []

    for spec in input_specs:
        name = spec["name"]
        value = str(execution_parameters.get(name) or "").strip()
        unresolved = _find_unresolved_placeholders(value)
        status = "available"
        error = ""
        if spec.get("required") and (not value or unresolved):
            status = "missing"
            error = f'Input "{name}" was not resolved before execution'
            missing.append(name)
        elif unresolved:
            status = "unresolved"
            error = f'Input "{name}" still contains unresolved placeholders'

        input_status[name] = {
            "name": name,
            "label": spec.get("label") or name,
            "required": bool(spec.get("required")),
            "status": status,
            "value": value if value and not unresolved else "",
            "error": error,
        }

    return missing, input_status


def _normalize_output_specs(outputs: Any) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    allowed_sources = {"page_text", "page_url", "page_title", "stdout", "stderr", "oracle_table"}

    for item in outputs or []:
        if not isinstance(item, dict):
            continue
        name = _normalize_param_name(str(item.get("name") or ""))
        source = str(item.get("source") or "page_text").strip().lower()
        pattern = str(item.get("pattern") or "").strip()
        group = item.get("group", 1)
        if not name or source not in allowed_sources:
            continue
        normalized.append(
            {
                "name": name,
                "source": source,
                "pattern": pattern,
                "group": group,
                "column": str(item.get("column") or "").strip(),
                "row": item.get("row", 0),
                "table_index": item.get("table_index", 0),
            }
        )
    return normalized


def _normalize_output_label(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip().lower()


def _flow_context_display_name(name: str) -> str:
    return re.sub(r"\s+", " ", str(name or "").replace("_", " ")).strip()


def _flow_context_label_candidates(spec: dict[str, Any]) -> list[str]:
    candidates: list[str] = []
    seen: set[str] = set()

    def _append(value: Any) -> None:
        cleaned = re.sub(r"\s+", " ", str(value or "").strip())
        if not cleaned:
            return
        normalized = cleaned.lower()
        if normalized in seen:
            return
        seen.add(normalized)
        candidates.append(cleaned)

    _append(spec.get("label"))
    for alias in _parse_flow_context_aliases(spec.get("aliases")):
        _append(alias)
    _append(spec.get("column"))
    _append(_flow_context_display_name(str(spec.get("name") or "")))
    return candidates


def _flow_context_label_match_score(candidate: str, target: str) -> int:
    left = _normalize_output_label(candidate)
    right = _normalize_output_label(target)
    if not left or not right:
        return 0
    if left == right:
        return 400
    if left in right or right in left:
        return 300 - abs(len(left) - len(right))
    left_tokens = set(left.split())
    right_tokens = set(right.split())
    overlap = left_tokens & right_tokens
    if not overlap:
        return 0
    return len(overlap) * 10


def _flow_context_best_header_match(headers: list[Any], spec: dict[str, Any]) -> tuple[int | None, str]:
    candidates = _flow_context_label_candidates(spec)
    best_index: int | None = None
    best_header = ""
    best_score = 0
    for index, header in enumerate(headers):
        header_text = str(header or "").strip()
        if not header_text:
            continue
        for candidate in candidates:
            score = _flow_context_label_match_score(header_text, candidate)
            if score <= best_score:
                continue
            best_index = index
            best_header = header_text
            best_score = score
    return best_index, best_header


def _extract_oracle_table_output(result: dict[str, Any], spec: dict[str, Any]) -> tuple[str | None, str | None]:
    tables = result.get("oracle_tables")
    if not isinstance(tables, list) or not tables:
        return None, f'{spec["name"]}: source "oracle_table" had no captured tables'

    try:
        table_index = max(0, int(spec.get("table_index", 0)))
    except Exception:
        table_index = 0
    if table_index >= len(tables):
        return None, f'{spec["name"]}: table index {table_index} was out of range'

    table = tables[table_index] if isinstance(tables[table_index], dict) else {}
    headers = table.get("headers") if isinstance(table.get("headers"), list) else []
    rows = table.get("rows") if isinstance(table.get("rows"), list) else []
    if not headers or not rows:
        return None, f'{spec["name"]}: selected table did not include headers and rows'

    target_column = _normalize_output_label(spec.get("column"))
    if not target_column:
        return None, f'{spec["name"]}: oracle_table output requires "column"'

    column_index = next(
        (index for index, header in enumerate(headers) if _normalize_output_label(header) == target_column),
        None,
    )
    if column_index is None:
        return None, f'{spec["name"]}: column "{spec.get("column")}" was not found in oracle table headers'

    try:
        row_index = max(0, int(spec.get("row", 0)))
    except Exception:
        row_index = 0
    if row_index >= len(rows):
        return None, f'{spec["name"]}: row index {row_index} was out of range'

    row = rows[row_index] if isinstance(rows[row_index], list) else []
    value = str(row[column_index] or "").strip() if column_index < len(row) else ""
    if not value:
        return None, f'{spec["name"]}: oracle table cell was empty at row {row_index}, column "{spec.get("column")}"'
    return value, None


def _extract_recording_outputs(
    result: dict[str, Any],
    outputs: Any,
) -> tuple[dict[str, str], list[str]]:
    extracted: dict[str, str] = {}
    errors: list[str] = []

    for spec in _normalize_output_specs(outputs):
        if spec["source"] == "oracle_table":
            extracted_value, error = _extract_oracle_table_output(result, spec)
            if error:
                errors.append(error)
                continue
            extracted[spec["name"]] = str(extracted_value or "").strip()
            continue

        source_value = str(result.get(spec["source"]) or "").strip()
        if not source_value:
            errors.append(f'{spec["name"]}: source "{spec["source"]}" was empty')
            continue

        pattern = str(spec.get("pattern") or "").strip()
        if not pattern:
            extracted_value = source_value
        else:
            try:
                match = re.search(pattern, source_value, flags=re.IGNORECASE | re.MULTILINE)
            except re.error as exc:
                errors.append(f'{spec["name"]}: invalid regex "{pattern}": {exc}')
                continue
            if not match:
                errors.append(f'{spec["name"]}: pattern "{pattern}" not found in {spec["source"]}')
                continue

            group = spec.get("group", 1)
            try:
                extracted_value = match.group(group)
            except Exception as exc:
                errors.append(f'{spec["name"]}: could not read regex group {group}: {exc}')
                continue

        value = str(extracted_value or "").strip()
        if not value:
            errors.append(f'{spec["name"]}: extracted value was empty')
            continue
        extracted[spec["name"]] = value

    return extracted, errors


def _flow_context_source_candidates(spec: dict[str, Any]) -> list[str]:
    source = str(spec.get("source") or "auto").strip().lower() or "auto"
    if source == "auto":
        return ["oracle_table", "page_semantics", "page_text", "page_title", "page_url", "stdout", "stderr"]
    if source == "ai":
        return []
    return [source]


def _flow_context_effective_label(spec: dict[str, Any]) -> str:
    candidates = _flow_context_label_candidates(spec)
    return candidates[0] if candidates else ""


def _flow_context_value_patterns(spec: dict[str, Any]) -> list[str]:
    value_type = str(spec.get("value_type") or "text").strip().lower()
    patterns: list[str] = []
    seen: set[str] = set()

    def _append(pattern: str) -> None:
        if not pattern or pattern in seen:
            return
        seen.add(pattern)
        patterns.append(pattern)

    explicit_pattern = str(spec.get("pattern") or "").strip()
    if explicit_pattern:
        _append(explicit_pattern)

    for label in _flow_context_label_candidates(spec):
        escaped = re.escape(label)
        if value_type in {"number", "integer", "id"}:
            _append(rf"{escaped}\s*(?:[:#-]\s*|\s+)(\d+)")
        else:
            _append(rf"{escaped}\s*[:#-]?\s*([^\n\r]+)")

    return patterns


def _normalize_flow_context_extracted_value(value: Any, spec: dict[str, Any]) -> str:
    text = re.sub(r"\s+", " ", str(value or "").strip())
    if not text:
        return ""

    explicit_pattern = str(spec.get("pattern") or "").strip()
    if explicit_pattern:
        try:
            match = re.search(explicit_pattern, text, flags=re.IGNORECASE | re.MULTILINE)
        except re.error:
            match = None
        if match:
            group = spec.get("group")
            if not isinstance(group, int):
                group = 1
            try:
                return str(match.group(group) or "").strip()
            except Exception:
                return ""

    value_type = str(spec.get("value_type") or "text").strip().lower()
    if value_type in {"number", "integer"}:
        match = re.search(r"\d+(?:\.\d+)?", text)
        return str(match.group(0) or "").strip() if match else ""
    if value_type == "id":
        match = re.search(r"[A-Za-z]+-\d+|\d+", text)
        return str(match.group(0) or "").strip() if match else ""
    return text


def _extract_flow_context_from_oracle_tables(
    result: dict[str, Any],
    spec: dict[str, Any],
) -> tuple[str | None, dict[str, Any]]:
    tables = result.get("oracle_tables")
    if not isinstance(tables, list) or not tables:
        return None, {"source": "oracle_table", "status": "miss", "detail": "No captured oracle tables"}

    if not _flow_context_label_candidates(spec):
        return None, {"source": "oracle_table", "status": "miss", "detail": "No output label or aliases were provided"}

    table_indexes: list[int]
    explicit_index = spec.get("table_index")
    if isinstance(explicit_index, int):
        table_indexes = [explicit_index]
    else:
        table_indexes = list(range(len(tables)))

    explicit_row_index = spec.get("row")
    row_index = explicit_row_index if isinstance(explicit_row_index, int) and explicit_row_index >= 0 else None

    for table_index in table_indexes:
        if table_index < 0 or table_index >= len(tables):
            continue
        table = tables[table_index] if isinstance(tables[table_index], dict) else {}
        headers = table.get("headers") if isinstance(table.get("headers"), list) else []
        rows = table.get("rows") if isinstance(table.get("rows"), list) else []
        if not headers or not rows:
            continue
        column_index, matched_header = _flow_context_best_header_match(headers, spec)
        if column_index is None:
            continue
        candidate_rows = [row_index] if row_index is not None else list(range(len(rows)))
        for candidate_row_index in candidate_rows:
            if candidate_row_index < 0 or candidate_row_index >= len(rows):
                continue
            row = rows[candidate_row_index] if isinstance(rows[candidate_row_index], list) else []
            value = str(row[column_index] or "").strip() if column_index < len(row) else ""
            if not value:
                continue
            detail = (
                f'table {table_index}, row {candidate_row_index}, column "{matched_header}"'
                if row_index is not None
                else f'table {table_index}, first non-empty row {candidate_row_index}, column "{matched_header}"'
            )
            return value, {
                "source": "oracle_table",
                "status": "matched",
                "detail": detail,
            }

    return None, {
        "source": "oracle_table",
        "status": "miss",
        "detail": f'No matching labelled value was found in captured oracle tables for "{_flow_context_effective_label(spec)}"',
    }


def _extract_flow_context_from_page_semantics(
    result: dict[str, Any],
    spec: dict[str, Any],
) -> tuple[str | None, dict[str, Any]]:
    semantics = result.get("page_semantics")
    if not isinstance(semantics, dict):
        return None, {"source": "page_semantics", "status": "miss", "detail": "No semantic snapshot captured"}

    label_values = semantics.get("label_values") if isinstance(semantics.get("label_values"), list) else []
    best_value = ""
    best_detail = ""
    best_score = 0

    for candidate in label_values:
        if not isinstance(candidate, dict):
            continue
        candidate_labels = [
            candidate.get("label"),
            candidate.get("title"),
            candidate.get("aria_label"),
            candidate.get("data_oj_field"),
        ]
        label_score = 0
        matched_label = ""
        for candidate_label in candidate_labels:
            candidate_text = str(candidate_label or "").strip()
            if not candidate_text:
                continue
            for requested_label in _flow_context_label_candidates(spec):
                score = _flow_context_label_match_score(candidate_text, requested_label)
                if score <= label_score:
                    continue
                label_score = score
                matched_label = candidate_text
        if label_score <= 0:
            continue
        normalized_value = _normalize_flow_context_extracted_value(candidate.get("value"), spec)
        if not normalized_value:
            continue
        if label_score <= best_score:
            continue
        best_score = label_score
        best_value = normalized_value
        best_detail = (
            f'label "{matched_label}" matched semantic field "{candidate.get("value")}"'
            if matched_label
            else "semantic label/value match"
        )

    if best_value:
        return best_value, {
            "source": "page_semantics",
            "status": "matched",
            "detail": best_detail,
        }

    dialogs = semantics.get("dialogs") if isinstance(semantics.get("dialogs"), list) else []
    for dialog in dialogs:
        if not isinstance(dialog, dict):
            continue
        dialog_text = "\n".join(
            part for part in (str(dialog.get("title") or "").strip(), str(dialog.get("text") or "").strip()) if part
        ).strip()
        if not dialog_text:
            continue
        value, attempt = _extract_flow_context_from_text_source(
            dialog_text,
            spec,
            source_name="page_semantics",
        )
        if value:
            attempt["detail"] = f'dialog {dialog.get("index", 0)}: {attempt.get("detail") or "matched"}'
            return value, attempt

    text_candidates = semantics.get("text_candidates") if isinstance(semantics.get("text_candidates"), list) else []
    for candidate in text_candidates:
        if not isinstance(candidate, dict):
            continue
        candidate_text = "\n".join(
            part
            for part in (
                str(candidate.get("text") or "").strip(),
                str(candidate.get("title") or "").strip(),
                str(candidate.get("aria_label") or "").strip(),
            )
            if part
        ).strip()
        if not candidate_text:
            continue
        value, attempt = _extract_flow_context_from_text_source(
            candidate_text,
            spec,
            source_name="page_semantics",
        )
        if value:
            attempt["detail"] = (
                f'{candidate.get("tag") or "element"}: {attempt.get("detail") or "matched"}'
            )
            return value, attempt

    return None, {
        "source": "page_semantics",
        "status": "miss",
        "detail": f'No matching labelled value was found in semantic snapshot for "{_flow_context_effective_label(spec)}"',
    }


def _extract_flow_context_from_text_source(
    text: str,
    spec: dict[str, Any],
    *,
    source_name: str,
) -> tuple[str | None, dict[str, Any]]:
    content = str(text or "").strip()
    if not content:
        return None, {"source": source_name, "status": "miss", "detail": f"{source_name} was empty"}

    patterns = _flow_context_value_patterns(spec)
    if not patterns:
        return None, {"source": source_name, "status": "miss", "detail": "No pattern or label available"}

    group = spec.get("group")
    if not isinstance(group, int):
        group = 1
    last_error = ""
    for pattern in patterns:
        try:
            match = re.search(pattern, content, flags=re.IGNORECASE | re.MULTILINE)
        except re.error as exc:
            last_error = f"Invalid regex {pattern}: {exc}"
            continue
        if not match:
            last_error = f'Pattern "{pattern}" not found'
            continue

        try:
            value = match.group(group)
        except Exception as exc:
            last_error = f"Could not read group {group}: {exc}"
            continue

        cleaned = str(value or "").strip()
        if not cleaned:
            last_error = "Matched value was empty"
            continue
        return cleaned, {"source": source_name, "status": "matched", "detail": f'Pattern "{pattern}" matched'}

    status = "error" if last_error.startswith("Invalid regex") else "miss"
    return None, {"source": source_name, "status": status, "detail": last_error or "No pattern matched"}


def _is_flow_context_ai_enabled() -> bool:
    raw = str(os.getenv("PTR_FLOW_CONTEXT_AI_EXTRACTION_ENABLED", "true")).strip().lower()
    return raw not in {"0", "false", "no", "off"}


def _get_flow_context_ai_model() -> str:
    return os.getenv("PTR_FLOW_CONTEXT_AI_MODEL", "gpt-4.1-mini").strip() or "gpt-4.1-mini"


def _build_flow_context_ai_request_payload(
    result: dict[str, Any],
    spec: dict[str, Any],
    *,
    model: str,
) -> tuple[dict[str, Any], str, str]:
    system_prompt = (
        "You extract a single business field from Playwright run diagnostics. "
        "Return concise JSON only."
    )
    user_prompt = textwrap.dedent(
        f"""
        Extract exactly one field from this Playwright execution snapshot and return JSON only.

        Return exactly:
        {{
          "value": string or null,
          "reason": string,
          "source": string,
          "confidence": "low" | "medium" | "high"
        }}

        Rules:
        - Extract only the requested field.
        - Prefer the most recent/current value visible in the provided snapshot.
        - If the value is missing or ambiguous, return "value": null and explain why.
        - Do not invent values.

        Requested field name: {spec.get("name")}
        Requested field label: {_flow_context_effective_label(spec)}
        Accepted labels and aliases: {json.dumps(_flow_context_label_candidates(spec), ensure_ascii=False)}
        Value type: {spec.get("value_type") or "text"}
        Additional hint: {spec.get("prompt") or "None"}

        Page title: {result.get("page_title") or ""}
        Page URL: {result.get("page_url") or ""}

        Oracle tables JSON:
        {json.dumps(result.get("oracle_tables") or [], ensure_ascii=False)}

        Semantic snapshot JSON:
        {json.dumps(result.get("page_semantics") or {}, ensure_ascii=False)}

        Page text:
        {_truncate_text(result.get("page_text"), max_chars=_MAX_FLOW_CONTEXT_PAGE_TEXT_CHARS) or "No page text captured."}
        """
    ).strip()

    request_payload = {
        "model": model,
        "input": [
            {
                "role": "system",
                "content": [{"type": "input_text", "text": system_prompt}],
            },
            {
                "role": "user",
                "content": [{"type": "input_text", "text": user_prompt}],
            },
        ],
        "text": {"format": {"type": "json_object"}},
        "max_output_tokens": 300,
    }
    return request_payload, system_prompt, user_prompt


def _call_openai_flow_context_extraction(
    result: dict[str, Any],
    spec: dict[str, Any],
) -> dict[str, Any]:
    api_key = str(os.getenv("OPENAI_API_KEY", "")).strip()
    model = _get_flow_context_ai_model()
    if not _is_flow_context_ai_enabled():
        return {
            "status": "skipped",
            "model": model,
            "reason": "Flow context AI extraction is disabled by PTR_FLOW_CONTEXT_AI_EXTRACTION_ENABLED.",
        }
    if not api_key:
        return {
            "status": "skipped",
            "model": model,
            "reason": "OPENAI_API_KEY is not configured.",
        }

    request_payload, system_prompt, user_prompt = _build_flow_context_ai_request_payload(
        result,
        spec,
        model=model,
    )
    request = Request(
        f"{_get_openai_base_url()}/responses",
        data=json.dumps(request_payload).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )

    try:
        with urlopen(request, timeout=45) as response:
            response_payload = json.loads(response.read().decode("utf-8"))
    except HTTPError as exc:
        details = exc.read().decode("utf-8", errors="replace")
        return {
            "status": "error",
            "model": model,
            "system_prompt": system_prompt,
            "user_prompt": user_prompt,
            "reason": f"OpenAI request failed ({exc.code}): {_summarize_openai_error(details)}",
        }
    except URLError as exc:
        return {
            "status": "error",
            "model": model,
            "system_prompt": system_prompt,
            "user_prompt": user_prompt,
            "reason": f"Unable to reach OpenAI API: {exc.reason}",
        }
    except Exception as exc:  # pragma: no cover - network/runtime path
        return {
            "status": "error",
            "model": model,
            "system_prompt": system_prompt,
            "user_prompt": user_prompt,
            "reason": f"Unexpected OpenAI failure: {exc}",
        }

    output_text = ""
    parsed: dict[str, Any] = {}
    try:
        output_text = _extract_response_output_text(response_payload)
        parsed = _parse_json_response(output_text)
    except Exception as exc:
        return {
            "status": "error",
            "model": model,
            "system_prompt": system_prompt,
            "user_prompt": user_prompt,
            "reason": f"Failed to parse OpenAI response: {exc}",
            "response_text": output_text,
        }

    usage = response_payload.get("usage") if isinstance(response_payload.get("usage"), dict) else {}
    return {
        "status": "success",
        "feature": "flow_context_extraction",
        "model": model,
        "system_prompt": system_prompt,
        "user_prompt": user_prompt,
        "response_text": output_text,
        "parsed_response": parsed,
        "usage": usage,
    }


def _extract_flow_context_outputs(
    result: dict[str, Any],
    flow_context_specs: list[dict[str, Any]],
) -> tuple[dict[str, str], list[dict[str, Any]], list[str]]:
    extracted: dict[str, str] = {}
    details: list[dict[str, Any]] = []
    errors: list[str] = []

    for spec in flow_context_specs:
        if spec.get("kind") != "output":
            continue

        attempts: list[dict[str, Any]] = []
        value: str | None = None
        matched_source = ""

        for source_name in _flow_context_source_candidates(spec):
            if source_name == "oracle_table":
                extracted_value, attempt = _extract_flow_context_from_oracle_tables(result, spec)
            elif source_name == "page_semantics":
                extracted_value, attempt = _extract_flow_context_from_page_semantics(result, spec)
            else:
                extracted_value, attempt = _extract_flow_context_from_text_source(
                    result.get(source_name),
                    spec,
                    source_name=source_name,
                )
            attempts.append(attempt)
            if extracted_value:
                value = extracted_value
                matched_source = source_name
                break

        ai_interaction: dict[str, Any] | None = None
        if value is None and spec.get("use_ai"):
            ai_interaction = _call_openai_flow_context_extraction(result, spec)
            ai_status = str(ai_interaction.get("status") or "unknown")
            ai_parsed = ai_interaction.get("parsed_response") if isinstance(ai_interaction.get("parsed_response"), dict) else {}
            ai_value = str(ai_parsed.get("value") or "").strip()
            attempts.append(
                {
                    "source": "ai",
                    "status": "matched" if ai_status == "success" and ai_value else ai_status,
                    "detail": ai_parsed.get("reason") or ai_interaction.get("reason") or "",
                }
            )
            if ai_status == "success" and ai_value:
                value = ai_value
                matched_source = "ai"

        status = "extracted" if value else "failed"
        error = ""
        if value:
            extracted[spec["name"]] = value
        elif spec.get("required"):
            error = f'Failed to extract required output "{spec["name"]}"'
            errors.append(error)

        details.append(
            {
                "name": spec["name"],
                "label": _flow_context_effective_label(spec),
                "required": bool(spec.get("required")),
                "status": status,
                "value": value or "",
                "source": matched_source,
                "attempts": attempts,
                "error": error,
                "ai_interaction": ai_interaction,
            }
        )

    return extracted, details, errors


def _truncate_text(value: Any, *, max_chars: int) -> str:
    text = str(value or "").strip()
    if len(text) <= max_chars:
        return text
    return f"{text[:max_chars]}... [truncated]"


def _is_ai_failure_summary_enabled() -> bool:
    raw = str(os.getenv("OPENAI_FAILURE_SUMMARY_ENABLED", "true")).strip().lower()
    return raw not in {"0", "false", "no", "off"}


def _get_openai_base_url() -> str:
    return os.getenv("OPENAI_BASE_URL", "https://api.openai.com/v1").rstrip("/")


def _get_openai_failure_summary_model() -> str:
    return os.getenv("OPENAI_FAILURE_SUMMARY_MODEL", "gpt-4.1-mini").strip() or "gpt-4.1-mini"


def _summarize_openai_error(raw_text: str) -> str:
    text = str(raw_text or "").strip()
    if len(text) <= _MAX_OPENAI_ERROR_CHARS:
        return text or "unknown error"
    return f"{text[:_MAX_OPENAI_ERROR_CHARS]}... [truncated]"


def _image_path_to_data_url(path: Path) -> str | None:
    if not path.exists() or not path.is_file():
        return None
    image_bytes = path.read_bytes()
    if not image_bytes or len(image_bytes) > _MAX_AI_IMAGE_BYTES:
        return None
    return f"data:image/png;base64,{base64.b64encode(image_bytes).decode('utf-8')}"


def _extract_response_output_text(payload: dict[str, Any]) -> str:
    direct = str(payload.get("output_text") or "").strip()
    if direct:
        return direct

    parts: list[str] = []
    for item in payload.get("output") or []:
        if not isinstance(item, dict):
            continue
        for content in item.get("content") or []:
            if not isinstance(content, dict):
                continue
            text = str(content.get("text") or "").strip()
            if text:
                parts.append(text)
    return "\n".join(parts).strip()


def _parse_json_response(text: str) -> dict[str, Any]:
    cleaned = text.strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned)
        cleaned = re.sub(r"\s*```$", "", cleaned)
    parsed = json.loads(cleaned)
    if not isinstance(parsed, dict):
        raise ValueError("AI response was not a JSON object.")
    return parsed


def _build_ai_failure_summary_prompt(result: dict[str, Any]) -> str:
    step_lines = []
    for step in list(result.get("step_artifacts") or []):
        step_lines.append(
            f"- Step {int(step.get('index') or 0)}: {str(step.get('action') or 'step')}"
        )
    step_text = "\n".join(step_lines) if step_lines else "- No captured steps available"

    return textwrap.dedent(
        f"""
        Analyze this failed Playwright test execution and return JSON only.

        Goals:
        - Identify the most likely root cause.
        - Point to the most likely failing step when possible.
        - Suggest the next debugging action or probable fix.
        - Be concise and evidence-based.
        - If evidence is weak, say so.

        Return exactly these keys:
        {{
          "headline": string,
          "summary": string,
          "failure_category": string,
          "suspected_step_index": integer or null,
          "confidence": "low" | "medium" | "high",
          "evidence": string[],
          "next_action": string
        }}

        Failure context:
        - Recording name: {result.get("recording_name") or result.get("file_key") or "unknown"}
        - File key: {result.get("file_key") or "unknown"}
        - Exit code: {result.get("exit_code")}
        - Page title: {result.get("page_title") or "unknown"}
        - Page URL: {result.get("page_url") or "unknown"}
        - Error: {_truncate_text(result.get("error"), max_chars=_MAX_AI_LOG_CHARS)}

        stderr:
        {_truncate_text(result.get("stderr"), max_chars=_MAX_AI_LOG_CHARS) or "No stderr captured."}

        stdout:
        {_truncate_text(result.get("stdout"), max_chars=_MAX_AI_LOG_CHARS) or "No stdout captured."}

        Captured steps:
        {step_text}
        """
    ).strip()


def _normalize_ai_failure_summary(payload: dict[str, Any], *, model: str) -> dict[str, Any]:
    evidence = payload.get("evidence")
    if not isinstance(evidence, list):
        evidence = []

    suspected_step_index = payload.get("suspected_step_index")
    if isinstance(suspected_step_index, bool):
        suspected_step_index = None
    elif suspected_step_index is not None:
        try:
            suspected_step_index = int(suspected_step_index)
        except (TypeError, ValueError):
            suspected_step_index = None

    confidence = str(payload.get("confidence") or "medium").strip().lower()
    if confidence not in {"low", "medium", "high"}:
        confidence = "medium"

    return {
        "status": "generated",
        "model": model,
        "headline": _truncate_text(payload.get("headline") or "AI failure summary", max_chars=160),
        "summary": _truncate_text(payload.get("summary") or "", max_chars=800),
        "failure_category": _truncate_text(payload.get("failure_category") or "unknown", max_chars=80),
        "suspected_step_index": suspected_step_index,
        "confidence": confidence,
        "evidence": [
            _truncate_text(item, max_chars=220) for item in evidence if str(item or "").strip()
        ][:4],
        "next_action": _truncate_text(payload.get("next_action") or "", max_chars=300),
    }


def _call_openai_failure_summary(
    result: dict[str, Any],
    *,
    failure_screenshot_path: Path | None,
    step_image_paths: list[Path],
) -> dict[str, Any]:
    api_key = str(os.getenv("OPENAI_API_KEY", "")).strip()
    if not _is_ai_failure_summary_enabled():
        return {
            "status": "skipped",
            "reason": "AI failure summaries are disabled by OPENAI_FAILURE_SUMMARY_ENABLED.",
        }
    if not api_key:
        return {
            "status": "skipped",
            "reason": "OPENAI_API_KEY is not configured.",
        }

    model = _get_openai_failure_summary_model()
    user_content: list[dict[str, Any]] = [
        {
            "type": "input_text",
            "text": _build_ai_failure_summary_prompt(result),
        }
    ]

    if failure_screenshot_path:
        failure_image = _image_path_to_data_url(failure_screenshot_path)
        if failure_image:
            user_content.append(
                {
                    "type": "input_image",
                    "image_url": failure_image,
                }
            )

    for image_path in step_image_paths[:_MAX_AI_STEP_IMAGES]:
        data_url = _image_path_to_data_url(image_path)
        if not data_url:
            continue
        user_content.append(
            {
                "type": "input_image",
                "image_url": data_url,
            }
        )

    request_payload = {
        "model": model,
        "input": [
            {
                "role": "system",
                "content": [
                    {
                        "type": "input_text",
                        "text": (
                            "You are a senior QA engineer diagnosing Playwright failures. "
                            "Use the logs, metadata, and screenshots to produce a concise, "
                            "actionable JSON summary."
                        ),
                    }
                ],
            },
            {
                "role": "user",
                "content": user_content,
            },
        ],
        "text": {
            "format": {
                "type": "json_object",
            }
        },
        "max_output_tokens": 500,
    }

    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    request = Request(
        f"{_get_openai_base_url()}/responses",
        data=json.dumps(request_payload).encode("utf-8"),
        headers=headers,
        method="POST",
    )

    try:
        with urlopen(request, timeout=45) as response:
            response_payload = json.loads(response.read().decode("utf-8"))
    except HTTPError as exc:
        details = exc.read().decode("utf-8", errors="replace")
        if exc.code == 401:
            return {
                "status": "error",
                "model": model,
                "reason": "AI configuration error: OPENAI_API_KEY is invalid or not configured correctly.",
            }
        return {
            "status": "error",
            "model": model,
            "reason": f"OpenAI request failed ({exc.code}): {_summarize_openai_error(details)}",
        }
    except URLError as exc:
        return {
            "status": "error",
            "model": model,
            "reason": f"Unable to reach OpenAI API: {exc.reason}",
        }
    except Exception as exc:  # pragma: no cover - network/runtime path
        return {
            "status": "error",
            "model": model,
            "reason": f"Unexpected OpenAI failure: {exc}",
        }

    try:
        output_text = _extract_response_output_text(response_payload)
        parsed = _parse_json_response(output_text)
        return _normalize_ai_failure_summary(parsed, model=model)
    except Exception as exc:
        return {
            "status": "error",
            "model": model,
            "reason": f"Failed to parse OpenAI response: {exc}",
        }


def _validate_python_playwright_script(script_text: str) -> None:
    trimmed = script_text.lstrip()

    obvious_js_markers = [
        "import {",
        "from '@playwright/test'",
        'from "@playwright/test"',
        "test(",
        "test.describe(",
        "=>",
        "const ",
        "let ",
        "await page.goto(",
    ]
    if any(marker in trimmed for marker in obvious_js_markers):
        raise ValueError(
            "Recording is not a Python Playwright script. "
            "playwright_test_runner currently supports Python recordings only."
        )

    python_markers = [
        "from playwright.async_api import",
        "from playwright.sync_api import",
        "async_playwright",
        "sync_playwright",
        "def run(",
        "def main(",
        "async def main(",
    ]
    if not any(marker in trimmed for marker in python_markers):
        raise ValueError("Recording does not look like a supported Python Playwright script.")


def _insert_after_future_imports(script_text: str, helper: str) -> str:
    lines = script_text.splitlines(keepends=True)
    idx = 0
    while idx < len(lines) and lines[idx].startswith("from __future__ import"):
        idx += 1
    prefix = "".join(lines[:idx])
    suffix = "".join(lines[idx:])
    return f"{prefix}{helper}\n\n{suffix}"


def _inject_runtime_helpers(script_text: str) -> str:
    """Backward-compatible shim for the retired embedded helper injector."""
    return _inject_runtime_helpers_v2(script_text)


def _inject_runtime_helpers_v2(script_text: str) -> str:
    helper_import = "from src.runtime.helpers_v2 import *"
    return _insert_after_future_imports(script_text, helper_import)


def _prepare_script_for_execution(script_text: str, parameters: dict[str, Any] | None = None) -> str:
    """AST-based script preparation pipeline.

    Uses the AST pipeline to:
      1. AST parse → structured action list (catches ALL locator patterns)
      2. Optimize → detect compound patterns (combobox+option, fill+enter, etc.)
      3. Generate → produce script where every action routes through _ptr_* helpers
      4. Import the clean runtime helper module used by prepared recordings

    Known coverage gaps fail fast with an explicit error so we do not silently
    replay unsupported raw actions through the old fallback-heavy runtime.
    """
    from src.runtime.parser import ParseCoverageError, parse_script
    from src.runtime.optimizer import optimize
    from src.runtime.script_generator import CoverageError, generate_full_script

    _validate_python_playwright_script(script_text)
    if parameters:
        script_text = _substitute_parameters(script_text, parameters)

    try:
        actions = parse_script(script_text)
        optimized = optimize(actions)
        generated_script = generate_full_script(optimized)
        logger.info(
            "AST pipeline: parsed %d actions, optimized to %d",
            len(actions),
            len(optimized),
        )
    except (ParseCoverageError, CoverageError) as exc:
        logger.warning(
            "AST pipeline rejected recording due to unsupported coverage: %s",
            exc,
        )
        raise RuntimeError(
            "Recording contains actions the AST runner does not safely support yet. "
            "Add helper coverage or adjust the recording before replaying it.\n"
            f"{exc}"
        ) from exc
    except Exception as exc:
        logger.exception(
            "AST-only preparation failed unexpectedly",
        )
        raise RuntimeError(
            "AST-only preparation failed unexpectedly. "
            "This recording is no longer eligible for legacy regex fallback.\n"
            f"{exc}"
        ) from exc

    # Import the clean runtime helper module instead of embedding the legacy
    # helper blob into every prepared script.
    return _inject_runtime_helpers_v2(generated_script)


def _prepare_script_via_ast(script_text: str, parameters: dict[str, Any] | None = None) -> str:
    return _prepare_script_for_execution(script_text, parameters)


def _read_failure_diagnostics(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        logger.exception("Failed to read diagnostics from %s", path)
        return {}


def _run_python_script(
    script_path: Path,
    working_dir: Path,
    *,
    timeout_seconds: int,
    env: dict[str, str],
) -> subprocess.CompletedProcess[str]:
    run_env = dict(env)
    run_env.setdefault("PYTHONUNBUFFERED", "1")
    python_bin = str(run_env.get("PLAYWRIGHT_TEST_PYTHON_BIN") or "python3").strip() or "python3"
    use_xvfb = _env_flag(run_env.get("PTR_USE_XVFB"), True)
    xvfb_bin = shutil.which("xvfb-run")
    command = [python_bin, str(script_path)]
    if use_xvfb and xvfb_bin:
        # Keep recorded scripts headed, but provide a virtual display from the
        # runner environment so we do not mutate the generated artifact.
        command = [
            xvfb_bin,
            "--auto-servernum",
            "--server-args=-screen 0 1440x900x24",
            *command,
        ]
    return subprocess.run(
        command,
        cwd=str(working_dir),
        capture_output=True,
        text=True,
        timeout=timeout_seconds,
        env=run_env,
    )


@tool()
async def expand_recordings_for_parameter_rows(recordings: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return await asyncio.to_thread(_expand_recordings_for_parameter_rows_data, recordings)


@tool()
async def load_resume_state_from_run(
    test_suite_id: str,
    previous_run_id: str,
    recordings: list[dict[str, Any]],
) -> dict[str, Any]:
    return await asyncio.to_thread(
        _load_resume_state_from_run_data,
        test_suite_id,
        previous_run_id,
        recordings,
    )


def _base_recording_result(recording: dict[str, Any]) -> dict[str, Any]:
    recording_id = str(recording.get("id") or "unknown")
    file_key = str(recording.get("file") or recording.get("recording_name") or "").strip()
    recording_name = str(recording.get("name") or "").strip() or file_key or recording_id

    return {
        "recording_id": recording_id,
        "recording_name": recording_name,
        "file_key": file_key,
        "parameter_row_index": recording.get("parameter_row_index"),
        "parameter_set_index": recording.get("parameter_set_index"),
        "status": "failed",
        "exit_code": -1,
        "duration_seconds": 0,
        "stdout": "",
        "stderr": "",
        "error": None,
        "page_url": None,
        "page_title": None,
        "page_text": None,
        "oracle_tables": [],
        "page_semantics": {},
        "screenshot_s3_key": None,
        "video_s3_key": None,
        "video_s3_keys": [],
        "step_artifacts": [],
        "ai_failure_summary": None,
        "parameters_file_key": None,
        "resolved_parameter_count": 0,
        "resolved_parameter_keys": [],
        "flow_context_file_key": None,
        "flow_context_specs": [],
        "flow_input_status": {},
        "flow_output_results": [],
        "extracted_outputs": {},
        "output_errors": [],
    }


@tool()
async def record_blocked_recording(
    test_suite_id: str,
    parent_run_id: str,
    recording: dict[str, Any],
    reason: str,
) -> dict[str, Any]:
    manifest_key = _manifest_key_for_recording(test_suite_id, parent_run_id, recording)
    result = _base_recording_result(recording)
    result["error"] = str(reason or "Recording was not executed because an upstream dependency failed.")
    result["blocked_by_dependency"] = True

    _storage_put_bytes(
        manifest_key,
        json.dumps(result, indent=2).encode("utf-8"),
        content_type="application/json",
    )
    result["result_s3_key"] = manifest_key
    return result


@tool()
async def execute_recording_script(
    recording: dict[str, Any],
    test_suite_id: str,
    parent_run_id: str,
) -> dict[str, Any]:
    file_key = str(recording.get("file") or recording.get("recording_name") or "").strip()
    artifact_identity = _recording_artifact_identity(recording)
    artifact_prefix = (
        f"playwright-test-results/{_safe_segment(test_suite_id)}/{_safe_segment(parent_run_id)}"
        f"/{_safe_segment(artifact_identity)}"
    )
    manifest_key = _manifest_key_for_recording(test_suite_id, parent_run_id, recording)

    result: dict[str, Any] = _base_recording_result(recording)

    start_time = time.time()

    if not file_key:
        result["error"] = "Recording file key is required."
        _storage_put_bytes(
            manifest_key,
            json.dumps(result, indent=2).encode("utf-8"),
            content_type="application/json",
        )
        result["result_s3_key"] = manifest_key
        return result

    try:
        raw_script_bytes = await asyncio.to_thread(_load_script_bytes, file_key)
        logger.info("Downloaded recording script for %s (%s bytes)", file_key, len(raw_script_bytes))

        # Auto-parameterise: extract hardcoded values as defaults and inject
        # {{placeholders}} in one pass — no manual script editing required.
        parameterised_script, default_params = _parameterise_script(raw_script_bytes.decode("utf-8"))
        logger.info("Auto-extracted %d default parameter(s) from script", len(default_params))

        # Merge order: script defaults → Excel file overrides → inline overrides.
        # Before execution we normalize the merged values into a JSON object and
        # substitute placeholders from that JSON payload.
        parameters: dict[str, str] = _normalize_parameter_values(default_params)
        parameters_file_key = str(recording.get("parameters_file_key") or "").strip() or None
        if not bool(recording.get("skip_parameters_file_load")):
            try:
                file_params, loaded_from = await asyncio.to_thread(_load_recording_parameters, recording, file_key)
                if file_params:
                    parameters.update(_normalize_parameter_values(file_params))
                    parameters_file_key = loaded_from
                    logger.info("Loaded %d parameter(s) from %s", len(file_params), loaded_from)
                elif loaded_from:
                    parameters_file_key = loaded_from
            except Exception as exc:
                parameters_file = str(recording.get("parameters_file") or "").strip()
                logger.warning("Failed to load parameters file %s: %s", parameters_file or file_key, exc)
        inline = _normalize_parameter_values(recording.get("parameters") or {})
        parameters.update(inline)

        flow_context_specs, flow_context_file_key = await asyncio.to_thread(
            _load_recording_flow_context_specs,
            recording,
            file_key,
        )
        normalized_flow_context_specs = _normalize_flow_context_specs(flow_context_specs)
        execution_parameters = _parameters_to_json_object(parameters)
        result["parameters_file_key"] = parameters_file_key
        result["resolved_parameter_count"] = len(execution_parameters)
        result["resolved_parameter_keys"] = sorted(execution_parameters)
        result["flow_context_file_key"] = flow_context_file_key
        result["flow_context_specs"] = normalized_flow_context_specs
        logger.info(
            "Resolved %d execution parameter(s) for %s: %s",
            len(execution_parameters),
            file_key,
            ", ".join(sorted(execution_parameters)),
        )

        missing_inputs, flow_input_status = _validate_flow_context_inputs(
            execution_parameters,
            normalized_flow_context_specs,
        )
        unresolved_parameters = _collect_unresolved_execution_parameters(execution_parameters)
        for name, placeholders in unresolved_parameters.items():
            existing = flow_input_status.get(name) if isinstance(flow_input_status, dict) else None
            flow_input_status[name] = {
                "name": name,
                "label": (existing or {}).get("label") or name,
                "required": bool((existing or {}).get("required")),
                "status": "unresolved",
                "value": "",
                "error": f'Parameter "{name}" still contains unresolved placeholders: {", ".join(placeholders)}',
            }
        result["flow_input_status"] = flow_input_status
        if missing_inputs or unresolved_parameters:
            error_parts: list[str] = []
            if missing_inputs:
                error_parts.append("required inputs: " + ", ".join(sorted(set(missing_inputs))))
            if unresolved_parameters:
                error_parts.append(
                    "unresolved parameters: "
                    + "; ".join(
                        f'{name} -> {", ".join(placeholders)}'
                        for name, placeholders in sorted(unresolved_parameters.items())
                    )
                )
            raise RuntimeError(
                "Required flow context values were not resolved before execution: "
                + " | ".join(error_parts)
            )

        prepared_script = _prepare_script_for_execution(
            parameterised_script,
            execution_parameters or None,
        )
    except Exception as exc:
        logger.exception("Failed to download or prepare recording script for %s", file_key)
        result["error"] = f"Failed to download or prepare recording script: {exc}"
        _storage_put_bytes(
            manifest_key,
            json.dumps(result, indent=2).encode("utf-8"),
            content_type="application/json",
        )
        result["result_s3_key"] = manifest_key
        return result

    with tempfile.TemporaryDirectory(prefix="playwright-test-runner-") as temp_dir:
        working_dir = Path(temp_dir)
        script_path = working_dir / f"{_safe_segment(Path(file_key).stem)}.py"
        script_path.write_text(prepared_script, encoding="utf-8")

        diagnostics_path = working_dir / "diagnostics.json"
        failure_screenshot_path = working_dir / "failure.png"
        step_artifacts_dir = working_dir / "steps"
        video_dir = working_dir / "video"

        env = _ensure_runner_pythonpath(_merge_runner_env_defaults(os.environ.copy()))
        env["PTR_DIAGNOSTICS_PATH"] = str(diagnostics_path)
        env["PTR_FAILURE_SCREENSHOT_PATH"] = str(failure_screenshot_path)
        env["PTR_EXECUTION_PARAMETERS_JSON"] = json.dumps(execution_parameters, sort_keys=True)
        experience_store_path = _default_experience_store_path()
        experience_store_path.parent.mkdir(parents=True, exist_ok=True)
        env["PTR_EXPERIENCE_STORE_PATH"] = str(experience_store_path)
        env.setdefault("PTR_EXPERIENCE_ENABLED", "true")
        env.setdefault("PTR_RUNNER_VERSION", "ptr-v2")
        if _env_flag(env.get("PTR_CAPTURE_STEPS"), True):
            step_artifacts_dir.mkdir(parents=True, exist_ok=True)
            env["PTR_STEP_ARTIFACTS_DIR"] = str(step_artifacts_dir)
        else:
            env.pop("PTR_STEP_ARTIFACTS_DIR", None)
        if _env_flag(env.get("PTR_RECORD_VIDEO"), False):
            video_dir.mkdir(parents=True, exist_ok=True)
            env["PTR_VIDEO_DIR"] = str(video_dir)
        else:
            env.pop("PTR_VIDEO_DIR", None)

        python_bin = str(env.get("PLAYWRIGHT_TEST_PYTHON_BIN") or "python3").strip() or "python3"

        try:
            completed = await asyncio.to_thread(
                _run_python_script,
                script_path,
                working_dir,
                timeout_seconds=900,
                env=env,
            )
        except subprocess.TimeoutExpired as exc:
            completed = subprocess.CompletedProcess(
                args=[python_bin, str(script_path)],
                returncode=124,
                stdout=exc.stdout or "",
                stderr=exc.stderr or f"Timed out after {exc.timeout}s",
            )
        except Exception as exc:
            logger.exception("Unexpected execution failure for %s", file_key)
            completed = subprocess.CompletedProcess(
                args=[python_bin, str(script_path)],
                returncode=1,
                stdout="",
                stderr=str(exc),
            )

        diagnostics = _read_failure_diagnostics(diagnostics_path)

        result["exit_code"] = int(completed.returncode)
        result["duration_seconds"] = round(time.time() - start_time, 3)
        result["stdout"] = completed.stdout or ""
        result["stderr"] = completed.stderr or ""
        result["status"] = "passed" if completed.returncode == 0 else "failed"
        result["error"] = None if completed.returncode == 0 else (completed.stderr or "Execution failed")
        result["page_url"] = diagnostics.get("page_url")
        result["page_title"] = diagnostics.get("page_title")
        result["page_text"] = diagnostics.get("page_text")
        result["oracle_tables"] = diagnostics.get("oracle_tables") or []
        result["page_semantics"] = diagnostics.get("page_semantics") or {}

        failure_local_path = diagnostics.get("failure_screenshot_path")
        failure_screenshot_path: Path | None = None
        if failure_local_path and Path(failure_local_path).exists():
            failure_screenshot_path = Path(failure_local_path)
            screenshot_key = f"{artifact_prefix}/failure.png"
            _storage_put_bytes(
                screenshot_key,
                failure_screenshot_path.read_bytes(),
                content_type="image/png",
            )
            result["screenshot_s3_key"] = screenshot_key

        step_artifacts: list[dict[str, Any]] = []
        step_image_paths: list[Path] = []
        for item in diagnostics.get("step_artifacts") or []:
            local_path = Path(str(item.get("local_path") or ""))
            if not local_path.exists():
                continue
            step_image_paths.append(local_path)
            screenshot_key = f"{artifact_prefix}/steps/{local_path.name}"
            _storage_put_bytes(
                screenshot_key,
                local_path.read_bytes(),
                content_type="image/png",
            )
            step_artifacts.append(
                {
                    "index": int(item.get("index") or 0),
                    "action": str(item.get("action") or "step"),
                    "screenshot_s3_key": screenshot_key,
                }
            )
        result["step_artifacts"] = step_artifacts
        result["action_log"] = diagnostics.get("action_log") or []

        # Oracle Fusion (and similar apps) open task pages in a new browser page,
        # so Playwright produces one .webm file per page. Upload all of them so
        # the caller can see the full recording across every page that was opened.
        video_files = sorted(video_dir.glob("*.webm"), key=lambda p: p.stat().st_mtime)
        video_s3_keys: list[str] = []
        for idx, vf in enumerate(video_files):
            vkey = f"{artifact_prefix}/recording_{idx}.webm"
            _storage_put_bytes(vkey, vf.read_bytes(), content_type="video/webm")
            video_s3_keys.append(vkey)
        if video_s3_keys:
            # Expose both the full list and a convenience key pointing to the last
            # (most recently opened) page — that is normally the one where the
            # failure occurred.
            result["video_s3_keys"] = video_s3_keys
            result["video_s3_key"] = video_s3_keys[-1]

        if result["status"] != "passed":
            result["ai_failure_summary"] = await asyncio.to_thread(
                _call_openai_failure_summary,
                result,
                failure_screenshot_path=failure_screenshot_path,
                step_image_paths=step_image_paths,
            )

        workbook_outputs, flow_output_results, flow_output_errors = _extract_flow_context_outputs(
            result,
            result.get("flow_context_specs") or [],
        )
        explicit_outputs, explicit_output_errors = _extract_recording_outputs(result, recording.get("outputs"))
        extracted_outputs = dict(workbook_outputs)
        extracted_outputs.update(explicit_outputs)
        result["flow_output_results"] = flow_output_results
        result["extracted_outputs"] = extracted_outputs
        result["output_errors"] = [*flow_output_errors, *explicit_output_errors]

    logger.info(
        "Finished recording %s with status=%s exit_code=%s duration=%ss",
        file_key,
        result["status"],
        result["exit_code"],
        result["duration_seconds"],
    )

    _storage_put_bytes(
        manifest_key,
        json.dumps(result, indent=2).encode("utf-8"),
        content_type="application/json",
    )
    result["result_s3_key"] = manifest_key
    return result


@tool()
async def generate_html_report(
    test_suite_id: str,
    parent_run_id: str,
    manifest_keys: dict[str, str],
    ordered_names: list[str],
) -> str:
    results: list[dict[str, Any]] = []
    for name in ordered_names:
        manifest_key = manifest_keys.get(name, "")
        if not manifest_key:
            results.append(
                {
                    "recording_name": name,
                    "status": "failed",
                    "duration_seconds": 0,
                    "stdout": "",
                    "stderr": "",
                    "error": "No manifest found for this recording.",
                    "page_url": None,
                    "page_title": None,
                    "screenshot_s3_key": None,
                    "step_artifacts": [],
                }
            )
            continue

        try:
            manifest = await asyncio.to_thread(_read_manifest, manifest_key)
            results.append(manifest)
        except Exception as exc:
            logger.exception("Failed to read manifest %s", manifest_key)
            results.append(
                {
                    "recording_name": name,
                    "status": "failed",
                    "duration_seconds": 0,
                    "stdout": "",
                    "stderr": "",
                    "error": f"Failed to load manifest: {exc}",
                    "page_url": None,
                    "page_title": None,
                    "screenshot_s3_key": None,
                    "step_artifacts": [],
                }
            )

    html_content = generate_html_report_content(
        test_suite_id=test_suite_id,
        parent_run_id=parent_run_id,
        results=results,
    )
    report_key = (
        f"playwright-test-results/{_safe_segment(test_suite_id)}/{_safe_segment(parent_run_id)}/report.html"
    )
    _storage_put_bytes(report_key, html_content.encode("utf-8"), content_type="text/html")
    return report_key
