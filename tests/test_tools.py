import subprocess
from pathlib import Path
from typing import Any

from src.tools.tools import (
    _collect_unresolved_execution_parameters,
    _default_experience_store_path,
    _derive_parameters_file_candidates,
    _extract_table_parameter_sets,
    _load_resume_state_from_run_data,
    _extract_flow_context_outputs,
    _extract_recording_outputs,
    _expand_recordings_for_parameter_rows_data,
    _manifest_key_for_recording,
    _load_runner_env_defaults,
    _merge_runner_env_defaults,
    _parameterise_script,
    _parse_flow_context_aliases,
    _parse_excel_flow_context_specs,
    _parameters_to_json_object,
    _parse_excel_parameter_sets,
    _parse_excel_parameters,
    _prepare_script_for_execution,
    _run_python_script,
    _split_storage_object_ref,
    _validate_flow_context_inputs,
)


def test_prepare_script_for_execution_rewrites_inline_search_result_clicks() -> None:
    script = """
from playwright.sync_api import Playwright, sync_playwright


def run(playwright: Playwright) -> None:
    browser = playwright.chromium.launch(headless=False)
    page = browser.new_page()
    page.get_by_role("textbox", name="Search for people to add as").fill("Fu Jiang")
    page.get_by_text("Fu Jiang").click()
    browser.close()


with sync_playwright() as playwright:
    run(playwright)
"""

    prepared = _prepare_script_for_execution(script)

    assert (
        "_ptr_tracked_action('search_and_select', 'Search for people to add as', "
        "_ptr_select_search_trigger_option, page.get_by_role('textbox', "
        "name='Search for people to add as'), page.get_by_text('Fu Jiang'), "
        "page, 'Search for people to add as', 'Fu Jiang', option_kind='text', fill_value='Fu Jiang')"
        in prepared
    )
    assert "_ptr_set_script_data({'tracked_action': 'search_and_select'" in prepared
    assert "_ptr_tracked_action('fill_textbox', 'Search for people to add as'" not in prepared
    assert "_ptr_tracked_action('click_text', 'Fu Jiang'" not in prepared


def test_prepare_script_for_execution_preserves_exact_search_result_clicks() -> None:
    script = """
from playwright.sync_api import Playwright, sync_playwright


def run(playwright: Playwright) -> None:
    browser = playwright.chromium.launch(headless=False)
    page = browser.new_page()
    page.get_by_role("textbox", name="Search for people to add as").fill("Fu")
    page.get_by_text("Wan Fu", exact=True).click()
    browser.close()


with sync_playwright() as playwright:
    run(playwright)
"""

    prepared = _prepare_script_for_execution(script)

    assert (
        "_ptr_tracked_action('search_and_select', 'Search for people to add as', "
        "_ptr_select_search_trigger_option, page.get_by_role('textbox', "
        "name='Search for people to add as'), page.get_by_text('Wan Fu', exact=True), "
        "page, 'Search for people to add as', 'Wan Fu', option_kind='text', "
        "option_exact=True, fill_value='Fu')"
        in prepared
    )


def test_manifest_key_for_recording_matches_result_artifact_layout() -> None:
    key = _manifest_key_for_recording(
        "hcm_first_3",
        "68128641-03ed-444a-8d36-0632ec3aa88a",
        {
            "id": "HCM_Move_To_Posting",
            "name": "HCM_Move_To_Posting",
            "file": "recordings/HCM_Move_To_Posting/HCM_Move_To_Posting.py",
        },
    )

    assert key == (
        "playwright-test-results/hcm_first_3/68128641-03ed-444a-8d36-0632ec3aa88a/"
        "recordings_HCM_Move_To_Posting_HCM_Move_To_Posting.py/manifest.json"
    )


def test_load_resume_state_from_run_data_starts_from_first_failed_recording(monkeypatch) -> None:
    recordings = [
        {
            "id": "HCM_Create_Requisition",
            "name": "HCM_Create_Requisition",
            "file": "recordings/HCM_Create_Requisition/HCM_Create_Requisition.py",
        },
        {
            "id": "HCM_Approve_Job_Requisition",
            "name": "HCM_Approve_Job_Requisition",
            "file": "recordings/HCM_Approve_Job_Requisition/HCM_Approve_Job_Requisition.py",
        },
        {
            "id": "HCM_Move_To_Posting",
            "name": "HCM_Move_To_Posting",
            "file": "recordings/HCM_Move_To_Posting/HCM_Move_To_Posting.py",
        },
    ]
    manifests = {
        _manifest_key_for_recording("hcm_first_3", "old-run", recordings[0]): {
            "status": "passed",
            "recording_name": "HCM_Create_Requisition",
            "extracted_outputs": {"requisition_id": "1008", "requisition_title": "Analyst"},
        },
        _manifest_key_for_recording("hcm_first_3", "old-run", recordings[1]): {
            "status": "passed",
            "recording_name": "HCM_Approve_Job_Requisition",
            "extracted_outputs": {},
        },
        _manifest_key_for_recording("hcm_first_3", "old-run", recordings[2]): {
            "status": "failed",
            "recording_name": "HCM_Move_To_Posting",
            "error": "textbox timeout",
        },
    }

    monkeypatch.setattr("src.tools.tools._read_manifest", lambda key: manifests[key])

    state = _load_resume_state_from_run_data("hcm_first_3", "old-run", recordings)

    assert state["resume_start_index"] == 2
    assert state["failed_recording_name"] == "HCM_Move_To_Posting"
    assert len(state["previous_results"]) == 2
    assert state["previous_results"][0]["extracted_outputs"]["requisition_id"] == "1008"


def test_load_resume_state_from_run_data_starts_from_first_missing_recording_after_passed_chain(
    monkeypatch,
) -> None:
    recordings = [
        {
            "id": "HCM_Create_Requisition",
            "name": "HCM_Create_Requisition",
            "file": "recordings/HCM_Create_Requisition/HCM_Create_Requisition.py",
        },
        {
            "id": "HCM_Approve_Job_Requisition",
            "name": "HCM_Approve_Job_Requisition",
            "file": "recordings/HCM_Approve_Job_Requisition/HCM_Approve_Job_Requisition.py",
        },
        {
            "id": "HCM_Move_To_Posting",
            "name": "HCM_Move_To_Posting",
            "file": "recordings/HCM_Move_To_Posting/HCM_Move_To_Posting.py",
        },
        {
            "id": "HCM_Add_Candidate_To_Requisition",
            "name": "HCM_Add_Candidate_To_Requisition",
            "file": "recordings/HCM_Add_Candidate_To_Requisition/HCM_Add_Candidate_To_Requisition.py",
        },
    ]
    manifests = {
        _manifest_key_for_recording("hcm_first_3", "old-run", recordings[0]): {
            "status": "passed",
            "recording_name": "HCM_Create_Requisition",
            "extracted_outputs": {"requisition_id": "1008", "requisition_title": "Analyst"},
        },
        _manifest_key_for_recording("hcm_first_3", "old-run", recordings[1]): {
            "status": "passed",
            "recording_name": "HCM_Approve_Job_Requisition",
            "extracted_outputs": {},
        },
        _manifest_key_for_recording("hcm_first_3", "old-run", recordings[2]): {
            "status": "passed",
            "recording_name": "HCM_Move_To_Posting",
            "extracted_outputs": {},
        },
    }

    def _fake_read_manifest(key: str) -> dict[str, Any]:
        if key not in manifests:
            raise FileNotFoundError(key)
        return manifests[key]

    monkeypatch.setattr("src.tools.tools._read_manifest", _fake_read_manifest)

    state = _load_resume_state_from_run_data("hcm_first_3", "old-run", recordings)

    assert state["resume_start_index"] == 3
    assert state["failed_recording_name"] == "HCM_Add_Candidate_To_Requisition"
    assert len(state["previous_results"]) == 3
    assert state["previous_results"][0]["extracted_outputs"]["requisition_id"] == "1008"


def test_prepare_script_for_execution_drops_premature_continue_before_reporting_relationship() -> None:
    script = """
from playwright.sync_api import Playwright, sync_playwright


def run(playwright: Playwright) -> None:
    browser = playwright.chromium.launch(headless=False)
    page = browser.new_page()
    page.get_by_role("combobox", name="Search for people to add as").click()
    page.get_by_role("textbox", name="Search for people to add as").fill("Fu")
    page.get_by_text("Wan Fu").click()
    page.get_by_role("button", name="Continue").click()
    page.get_by_role("combobox", name="Reporting Relationship").click()
    page.get_by_role("option", name="Project Manager").click()
    browser.close()


with sync_playwright() as playwright:
    run(playwright)
"""

    prepared = _prepare_script_for_execution(script)

    assert (
        "_ptr_tracked_action('search_and_select', 'Search for people to add as', "
        "_ptr_select_search_trigger_option, page.get_by_role('textbox', "
        "name='Search for people to add as'), page.get_by_text('Wan Fu'), "
        "page, 'Search for people to add as', 'Wan Fu', option_kind='text', fill_value='Fu')"
        in prepared
    )
    assert (
        "_ptr_tracked_action('select_combobox', 'Reporting Relationship', "
        "_ptr_select_combobox_option, page.get_by_role('combobox', "
        "name='Reporting Relationship'), page.get_by_role('option', "
        "name='Project Manager'), page, 'Reporting Relationship', "
        "'Project Manager')"
        in prepared
    )
    assert "_ptr_tracked_action('click_combobox', 'Search for people to add as'" not in prepared
    assert "_ptr_tracked_action('navigation_button', 'Continue'" not in prepared


def test_parameterise_script_does_not_reuse_stale_textbox_context_for_calendar_gridcell() -> None:
    script = """
page.get_by_role("textbox", name="Password").fill("1234567890")
page.get_by_role("textbox", name="Password").press("Enter")
page.get_by_title("Select Date.").click()
page.get_by_role("gridcell", name="9").click()
"""

    parameterised_script, default_params = _parameterise_script(script)
    gridcell_line = next(
        line for line in parameterised_script.splitlines() if 'get_by_role("gridcell"' in line
    )

    assert default_params["password"] == "1234567890"
    assert 'name="9"' in gridcell_line
    assert "{{password}}" not in gridcell_line


def test_parameterise_script_keeps_immediate_combobox_context_for_gridcell_selection() -> None:
    script = """
page.get_by_role("combobox", name="What's the way to change the assignment?").click()
page.get_by_role("gridcell", name="Temporary Assignment").click()
"""

    parameterised_script, default_params = _parameterise_script(script)

    assert default_params["what_s_the_way_to_change_the_assignment"] == "Temporary Assignment"
    assert (
        'page.get_by_role("gridcell", name="{{what_s_the_way_to_change_the_assignment}}").click()'
        in parameterised_script
    )


def test_parameterise_script_does_not_treat_existing_placeholders_as_defaults() -> None:
    script = """
page.goto("{{url}}")
page.get_by_role("textbox", name="Username").fill("{{username}}")
"""

    parameterised_script, default_params = _parameterise_script(script)

    assert 'page.goto("{{url}}")' in parameterised_script
    assert 'page.get_by_role("textbox", name="Username").fill("{{username}}")' in parameterised_script
    assert default_params == {}


def test_parameters_to_json_object_resolves_references_between_parameters() -> None:
    resolved = _parameters_to_json_object(
        {
            "requisition_id": "REQ-10025",
            "search_value": "{{requisition_id}}",
            "approval_lookup": "Request {{requisition_id}}",
        }
    )

    assert resolved["requisition_id"] == "REQ-10025"
    assert resolved["search_value"] == "REQ-10025"
    assert resolved["approval_lookup"] == "Request REQ-10025"


def test_extract_recording_outputs_reads_regex_from_page_text() -> None:
    extracted, errors = _extract_recording_outputs(
        {
            "page_text": "Requisition REQ-10025 was created successfully.",
            "page_url": "",
            "page_title": "",
            "stdout": "",
            "stderr": "",
        },
        [
            {
                "name": "requisition_id",
                "source": "page_text",
                "pattern": r"Requisition\s+(REQ-\d+)",
            }
        ],
    )

    assert extracted == {"requisition_id": "REQ-10025"}
    assert errors == []


def test_extract_recording_outputs_reads_oracle_table_first_row_cell() -> None:
    extracted, errors = _extract_recording_outputs(
        {
            "page_text": "",
            "page_url": "",
            "page_title": "",
            "stdout": "",
            "stderr": "",
            "oracle_tables": [
                {
                    "headers": ["Requisition Title", "Requisition Number", "Requisition Status"],
                    "rows": [
                        ["Analyst", "1003", "Approval - Pending"],
                        ["Analyst", "1002", "Approval - Pending"],
                    ],
                }
            ],
        },
        [
            {
                "name": "requisition_id",
                "source": "oracle_table",
                "column": "Requisition Number",
                "row": 0,
                "table_index": 0,
            }
        ],
    )

    assert extracted == {"requisition_id": "1003"}
    assert errors == []


def test_load_runner_env_defaults_parses_exported_values(tmp_path) -> None:
    config_path = tmp_path / "configs.txt"
    config_path.write_text(
        "\n".join(
            [
                "# comment",
                "export PTR_CAPTURE_STEPS=true",
                "PTR_POST_CLICK_WAIT_MS=250",
                'export PTR_GREETING="hello world"',
                "invalid line",
            ]
        ),
        encoding="utf-8",
    )

    defaults = _load_runner_env_defaults(config_path)

    assert defaults == {
        "PTR_CAPTURE_STEPS": "true",
        "PTR_POST_CLICK_WAIT_MS": "250",
        "PTR_GREETING": "hello world",
    }


def test_merge_runner_env_defaults_keeps_explicit_env_values(tmp_path) -> None:
    config_path = tmp_path / "configs.txt"
    config_path.write_text(
        "\n".join(
            [
                "export PTR_RECORD_VIDEO=false",
                "export PTR_POST_CLICK_WAIT_MS=250",
            ]
        ),
        encoding="utf-8",
    )

    merged = _merge_runner_env_defaults(
        {
            "PTR_POST_CLICK_WAIT_MS": "900",
            "CUSTOM_VALUE": "present",
        },
        config_path=config_path,
    )

    assert merged["PTR_RECORD_VIDEO"] == "false"
    assert merged["PTR_POST_CLICK_WAIT_MS"] == "900"
    assert merged["CUSTOM_VALUE"] == "present"


def test_run_python_script_uses_xvfb_when_available(monkeypatch, tmp_path: Path) -> None:
    captured: dict[str, Any] = {}

    def fake_run(*args: Any, **kwargs: Any) -> subprocess.CompletedProcess[str]:
        captured["args"] = args
        captured["kwargs"] = kwargs
        return subprocess.CompletedProcess(args[0], 0, stdout="", stderr="")

    monkeypatch.setattr("src.tools.tools.shutil.which", lambda name: "/usr/bin/xvfb-run")
    monkeypatch.setattr("src.tools.tools.subprocess.run", fake_run)

    script_path = tmp_path / "recording.py"
    script_path.write_text("print('ok')", encoding="utf-8")

    result = _run_python_script(
        script_path,
        tmp_path,
        timeout_seconds=30,
        env={},
    )

    assert result.returncode == 0
    assert captured["args"][0] == [
        "/usr/bin/xvfb-run",
        "--auto-servernum",
        "--server-args=-screen 0 1440x900x24",
        "python3",
        str(script_path),
    ]
    assert captured["kwargs"]["cwd"] == str(tmp_path)


def test_run_python_script_skips_xvfb_when_disabled(monkeypatch, tmp_path: Path) -> None:
    captured: dict[str, Any] = {}

    def fake_run(*args: Any, **kwargs: Any) -> subprocess.CompletedProcess[str]:
        captured["args"] = args
        captured["kwargs"] = kwargs
        return subprocess.CompletedProcess(args[0], 0, stdout="", stderr="")

    monkeypatch.setattr("src.tools.tools.shutil.which", lambda name: "/usr/bin/xvfb-run")
    monkeypatch.setattr("src.tools.tools.subprocess.run", fake_run)

    script_path = tmp_path / "recording.py"
    script_path.write_text("print('ok')", encoding="utf-8")

    _run_python_script(
        script_path,
        tmp_path,
        timeout_seconds=30,
        env={"PTR_USE_XVFB": "false"},
    )

    assert captured["args"][0] == ["python3", str(script_path)]


def test_default_experience_store_path_uses_runner_data_dir(tmp_path) -> None:
    path = _default_experience_store_path(tmp_path)

    assert path == tmp_path.resolve() / ".runner_data" / "experience.jsonl"


def test_parse_excel_parameters_supports_headerless_sheet() -> None:
    from io import BytesIO

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["url", "https://example.com/login"])
    ws.append(["username", "demo@example.com"])
    ws.append(["click_save", "ignored"])

    buffer = BytesIO()
    wb.save(buffer)
    wb.close()

    params = _parse_excel_parameters(buffer.getvalue())

    assert params == {
        "url": "https://example.com/login",
        "username": "demo@example.com",
    }


def test_parse_excel_parameters_supports_horizontal_header_value_sheet() -> None:
    from io import BytesIO

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            "startUrl",
            "Username",
            "Password",
            "Business Unit",
            "Receipt Number",
        ]
    )
    ws.append(
        [
            "https://iamoqy-test.fa.ocs.oraclecloud.com/",
            "svc",
            "Calfus@123",
            "Test Solutions",
            "RN-465346",
        ]
    )

    buffer = BytesIO()
    wb.save(buffer)
    wb.close()

    params = _parse_excel_parameters(buffer.getvalue())

    assert params == {
        "url": "https://iamoqy-test.fa.ocs.oraclecloud.com/",
        "username": "svc",
        "password": "Calfus@123",
        "business_unit": "Test Solutions",
        "receipt_number": "RN-465346",
    }


def test_parse_excel_parameter_sets_supports_multiple_horizontal_data_rows() -> None:
    from io import BytesIO

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            "startUrl",
            "Username",
            "Password",
            "Receipt Number",
            "Entered Amount",
        ]
    )
    ws.append(
        [
            "https://iamoqy-test.fa.ocs.oraclecloud.com/",
            "svc",
            "Calfus@123",
            "RN-46534434",
            "55",
        ]
    )
    ws.append(
        [
            "https://iamoqy-test.fa.ocs.oraclecloud.com/",
            "svc",
            "Calfus@124",
            "RN-46534734",
            "56",
        ]
    )

    buffer = BytesIO()
    wb.save(buffer)
    wb.close()

    parameter_sets = _parse_excel_parameter_sets(buffer.getvalue())

    assert parameter_sets == [
        {
            "row_index": 2,
            "values": {
                "url": "https://iamoqy-test.fa.ocs.oraclecloud.com/",
                "username": "svc",
                "password": "Calfus@123",
                "receipt_number": "RN-46534434",
                "entered_amount": "55",
            },
        },
        {
            "row_index": 3,
            "values": {
                "url": "https://iamoqy-test.fa.ocs.oraclecloud.com/",
                "username": "svc",
                "password": "Calfus@124",
                "receipt_number": "RN-46534734",
                "entered_amount": "56",
            },
        },
    ]


def test_parse_excel_parameter_sets_prefers_parameter_value_sheet_over_horizontal_rows() -> None:
    from io import BytesIO

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["Parameter", "Value", "Description", "delay_before_ms"])
    ws.append(["startUrl", "https://iamoqy-test.fa.ocs.oraclecloud.com/", "Base URL", ""])
    ws.append(["Username", "svc", "Login user", ""])
    ws.append(["Password", "Calfus@123", "Login password", ""])
    ws.append(["Search by Name Business", "Apple, Inc.", "Worker search", "500"])

    buffer = BytesIO()
    wb.save(buffer)
    wb.close()

    parameter_sets = _parse_excel_parameter_sets(buffer.getvalue())

    assert parameter_sets == [
        {
            "row_index": 2,
            "values": {
                "url": "https://iamoqy-test.fa.ocs.oraclecloud.com/",
                "username": "svc",
                "password": "Calfus@123",
                "search_by_name_business": "Apple, Inc.",
            },
        }
    ]


def test_parse_excel_flow_context_specs_reads_second_sheet_configuration() -> None:
    from io import BytesIO

    from openpyxl import Workbook

    wb = Workbook()
    params = wb.active
    params.title = "Params"
    params.append(["Parameter", "Value"])
    params.append(["url", "https://example.com"])

    flow_context = wb.create_sheet("Flow Context")
    flow_context.append(
        ["kind", "name", "label", "aliases", "source", "required", "use_ai", "prompt", "value_type"]
    )
    flow_context.append(
        [
            "output",
            "requisition_id",
            "Requisition Number",
            "Req Number|Job Requisition Number",
            "auto",
            "yes",
            "yes",
            "Extract the created requisition number",
            "number",
        ]
    )
    flow_context.append(["input", "search_value", "Search Value", "", "", "yes", "no", "", ""])

    buffer = BytesIO()
    wb.save(buffer)
    wb.close()

    specs = _parse_excel_flow_context_specs(buffer.getvalue())

    assert specs == [
        {
            "row_index": 2,
            "kind": "output",
            "name": "requisition_id",
            "label": "Requisition Number",
            "aliases": ["Req Number", "Job Requisition Number"],
            "source": "auto",
            "pattern": "",
            "group": 1,
            "column": "",
            "row": None,
            "table_index": None,
            "required": True,
            "prompt": "Extract the created requisition number",
            "value_type": "number",
            "use_ai": True,
        },
        {
            "row_index": 3,
            "kind": "input",
            "name": "search_value",
            "label": "Search Value",
            "aliases": [],
            "source": "auto",
            "pattern": "",
            "group": 1,
            "column": "",
            "row": None,
            "table_index": None,
            "required": True,
            "prompt": "",
            "value_type": "text",
            "use_ai": False,
        },
    ]


def test_parse_flow_context_aliases_accepts_existing_lists_without_stringifying() -> None:
    assert _parse_flow_context_aliases(["Req Number", "Job Requisition Number"]) == [
        "Req Number",
        "Job Requisition Number",
    ]


def test_extract_table_parameter_sets_ignores_misaligned_header_like_rows() -> None:
    parameter_sets = _extract_table_parameter_sets(
        [
            (
                "startUrl",
                "Username",
                "Password",
                "requisition_id",
                "requisition_title",
                "posting_schedule",
            ),
            (
                "https://fa-eqha-dev17-saasfademo1.ds-fa.oraclepdemos.com/",
                "HCM_IMPL",
                "En6*8fj%",
                "{{requisition_id}}",
                "{{requisition_title}}",
                "Post Now",
            ),
            ("requisition_id", "{{requisition_id}}", "", "", "", ""),
            ("requisition_title", "{{requisition_title}}", "", "", "", ""),
        ]
    )

    assert parameter_sets == [
        {
            "row_index": 2,
            "values": {
                "url": "https://fa-eqha-dev17-saasfademo1.ds-fa.oraclepdemos.com/",
                "username": "HCM_IMPL",
                "password": "En6*8fj%",
                "requisition_id": "{{requisition_id}}",
                "requisition_title": "{{requisition_title}}",
                "posting_schedule": "Post Now",
            },
        }
    ]


def test_expand_recordings_for_parameter_rows_data_fans_out_multiple_excel_rows(monkeypatch) -> None:
    def _fake_load_recording_parameter_sets(recording, file_key):
        assert recording["file"] == "recordings/demo.py"
        assert file_key == "recordings/demo.py"
        return (
            [
                {
                    "row_index": 2,
                    "values": {
                        "username": "svc",
                        "password": "Calfus@123",
                        "receipt_number": "RN-46534434",
                    },
                },
                {
                    "row_index": 3,
                    "values": {
                        "username": "svc",
                        "password": "Calfus@124",
                        "receipt_number": "RN-46534734",
                    },
                },
            ],
            "recordings/demo_params.xlsx",
        )

    monkeypatch.setattr(
        "src.tools.tools._load_recording_parameter_sets",
        _fake_load_recording_parameter_sets,
    )

    expanded = _expand_recordings_for_parameter_rows_data(
        [
            {
                "id": "rec-1",
                "name": "fake_2",
                "file": "recordings/demo.py",
                "parameters": {
                    "entered_amount": "55",
                },
            }
        ]
    )

    assert len(expanded) == 2
    assert expanded[0]["id"] == "rec-1-row-2"
    assert expanded[0]["name"] == "fake_2 [row 2]"
    assert expanded[0]["parameters"] == {
        "username": "svc",
        "password": "Calfus@123",
        "receipt_number": "RN-46534434",
        "entered_amount": "55",
    }
    assert expanded[0]["parameters_file_key"] == "recordings/demo_params.xlsx"
    assert expanded[0]["parameter_set_index"] == 1
    assert expanded[0]["parameter_row_index"] == 2
    assert expanded[0]["skip_parameters_file_load"] is True

    assert expanded[1]["id"] == "rec-1-row-3"
    assert expanded[1]["name"] == "fake_2 [row 3]"
    assert expanded[1]["parameters"] == {
        "username": "svc",
        "password": "Calfus@124",
        "receipt_number": "RN-46534734",
        "entered_amount": "55",
    }
    assert expanded[1]["parameters_file_key"] == "recordings/demo_params.xlsx"
    assert expanded[1]["parameter_set_index"] == 2
    assert expanded[1]["parameter_row_index"] == 3
    assert expanded[1]["skip_parameters_file_load"] is True


def test_parameters_to_json_object_normalizes_inline_parameter_keys() -> None:
    params = _parameters_to_json_object(
        {
            "startUrl": "https://iamoqy-test.fa.ocs.oraclecloud.com/",
            "Username": "svc",
            "Receipt Number": "RN-465346",
            "ignored_blank": "   ",
            "none_value": None,
        }
    )

    assert params == {
        "url": "https://iamoqy-test.fa.ocs.oraclecloud.com/",
        "username": "svc",
        "receipt_number": "RN-465346",
    }


def test_split_storage_object_ref_strips_current_bucket_prefix(monkeypatch) -> None:
    monkeypatch.setenv("STORAGE_ACTIVITIES_BUCKET", "local-dev-bucket")

    bucket_name, object_key = _split_storage_object_ref(
        "local-dev-bucket/recordings/8279897e-21b5-4781-ab82-fd4bd0095355/fake_2_params.xlsx"
    )

    assert bucket_name == "local-dev-bucket"
    assert object_key == "recordings/8279897e-21b5-4781-ab82-fd4bd0095355/fake_2_params.xlsx"


def test_derive_parameters_file_candidates_uses_sibling_params_file(monkeypatch) -> None:
    monkeypatch.setenv("STORAGE_ACTIVITIES_BUCKET", "local-dev-bucket")

    candidates = _derive_parameters_file_candidates(
        "local-dev-bucket/recordings/8279897e-21b5-4781-ab82-fd4bd0095355/fake_2.py"
    )

    assert candidates == [
        "local-dev-bucket/recordings/8279897e-21b5-4781-ab82-fd4bd0095355/fake_2_params.xlsx",
        "recordings/8279897e-21b5-4781-ab82-fd4bd0095355/fake_2_params.xlsx",
        "local-dev-bucket/recordings/8279897e-21b5-4781-ab82-fd4bd0095355/fake_2_params.csv",
        "recordings/8279897e-21b5-4781-ab82-fd4bd0095355/fake_2_params.csv",
    ]


def test_validate_flow_context_inputs_flags_missing_required_values() -> None:
    missing, status = _validate_flow_context_inputs(
        {"search_value": ""},
        [
            {
                "kind": "input",
                "name": "search_value",
                "label": "Search Value",
                "required": True,
            }
        ],
    )

    assert missing == ["search_value"]
    assert status["search_value"]["status"] == "missing"


def test_collect_unresolved_execution_parameters_detects_raw_placeholders() -> None:
    unresolved = _collect_unresolved_execution_parameters(
        {
            "search_value": "{{requisition_id}}",
            "title": "Analyst",
            "approval_label": "Request {{requisition_id}} pending",
        }
    )

    assert unresolved == {
        "search_value": ["requisition_id"],
        "approval_label": ["requisition_id"],
    }


def test_extract_flow_context_outputs_reads_first_matching_table_value_without_row_or_column_hints() -> None:
    extracted, details, errors = _extract_flow_context_outputs(
        {
            "page_text": "",
            "page_url": "",
            "page_title": "",
            "stdout": "",
            "stderr": "",
            "oracle_tables": [
                {
                    "headers": ["Requisition Title", "Requisition Number", "Requisition Status"],
                    "rows": [["Analyst", "1003", "Approval - Pending"], ["Senior Analyst", "1002", "Draft"]],
                }
            ],
        },
        [
            {
                "kind": "output",
                "name": "requisition_id",
                "label": "Requisition Number",
                "aliases": ["Req Number"],
                "source": "auto",
                "required": True,
                "use_ai": True,
                "value_type": "number",
            }
        ],
    )

    assert extracted == {"requisition_id": "1003"}
    assert errors == []
    assert details[0]["status"] == "extracted"
    assert details[0]["source"] == "oracle_table"
    assert details[0]["attempts"][0]["status"] == "matched"
    assert details[0]["attempts"][0]["detail"] == 'table 0, first non-empty row 0, column "Requisition Number"'


def test_extract_flow_context_outputs_reads_labelled_value_from_page_text_without_location_hint() -> None:
    extracted, details, errors = _extract_flow_context_outputs(
        {
            "page_text": "Review complete. Requisition Number 1007 was created for Analyst.",
            "page_url": "",
            "page_title": "",
            "stdout": "",
            "stderr": "",
            "oracle_tables": [],
        },
        [
            {
                "kind": "output",
                "name": "requisition_id",
                "label": "Requisition Number",
                "source": "auto",
                "required": True,
                "use_ai": True,
                "value_type": "number",
            }
        ],
    )

    assert extracted == {"requisition_id": "1007"}
    assert errors == []
    assert details[0]["source"] == "page_text"
    assert details[0]["attempts"][1]["source"] == "page_semantics"
    assert details[0]["attempts"][1]["status"] == "miss"
    assert details[0]["attempts"][2]["source"] == "page_text"
    assert details[0]["attempts"][2]["status"] == "matched"


def test_extract_flow_context_outputs_reads_labelled_value_from_page_semantics_before_page_text() -> None:
    extracted, details, errors = _extract_flow_context_outputs(
        {
            "page_text": "Review complete. Requisition Number 1007 was created for Analyst.",
            "page_url": "",
            "page_title": "",
            "stdout": "",
            "stderr": "",
            "oracle_tables": [],
            "page_semantics": {
                "label_values": [
                    {
                        "label": "Requisition Number",
                        "value": "1008",
                        "tag": "span",
                        "role": "",
                        "id": "req-number",
                        "title": "",
                        "aria_label": "",
                        "data_oj_field": "",
                    }
                ],
                "text_candidates": [],
                "dialogs": [],
            },
        },
        [
            {
                "kind": "output",
                "name": "requisition_id",
                "label": "Requisition Number",
                "aliases": ["Req Number"],
                "source": "auto",
                "required": True,
                "use_ai": True,
                "value_type": "number",
            }
        ],
    )

    assert extracted == {"requisition_id": "1008"}
    assert errors == []
    assert details[0]["source"] == "page_semantics"
    assert details[0]["attempts"][1]["source"] == "page_semantics"
    assert details[0]["attempts"][1]["status"] == "matched"


def test_extract_flow_context_outputs_falls_back_to_ai_when_needed(monkeypatch) -> None:
    def _fake_ai_extract(result, spec):
        return {
            "status": "success",
            "feature": "flow_context_extraction",
            "model": "gpt-4.1-mini",
            "system_prompt": "system",
            "user_prompt": "user",
            "response_text": '{"value":"1004","reason":"Found in confirmation banner","source":"page_text","confidence":"high"}',
            "parsed_response": {
                "value": "1004",
                "reason": "Found in confirmation banner",
                "source": "page_text",
                "confidence": "high",
            },
            "usage": {"input_tokens": 100, "output_tokens": 20, "total_tokens": 120},
        }

    monkeypatch.setattr("src.tools.tools._call_openai_flow_context_extraction", _fake_ai_extract)

    extracted, details, errors = _extract_flow_context_outputs(
        {
            "page_text": "",
            "page_url": "",
            "page_title": "",
            "stdout": "",
            "stderr": "",
            "oracle_tables": [],
        },
        [
            {
                "kind": "output",
                "name": "requisition_id",
                "label": "Requisition Number",
                "source": "auto",
                "required": True,
                "use_ai": True,
            }
        ],
    )

    assert extracted == {"requisition_id": "1004"}
    assert errors == []
    assert details[0]["source"] == "ai"
    assert details[0]["attempts"][-1]["source"] == "ai"
    assert details[0]["ai_interaction"]["model"] == "gpt-4.1-mini"
